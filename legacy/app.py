from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, session, g
from flask_sqlalchemy import SQLAlchemy
from datetime import datetime, date
import json
from translations import get_t

import os as _os

# Load .env.local from project root (one level up from legacy/)
try:
    from dotenv import load_dotenv
    _ENV_FILE = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..', '.env.local')
    load_dotenv(_ENV_FILE)
except ImportError:
    pass  # dotenv optional in production

app = Flask(__name__)
_DB_PATH = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), 'instance', 'accounting.db')
_os.makedirs(_os.path.dirname(_DB_PATH), exist_ok=True)
app.config['SQLALCHEMY_DATABASE_URI'] = f'sqlite:///{_DB_PATH}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = _os.environ.get('APP_SECRET_KEY', 'tumenaccounting2026')
db = SQLAlchemy(app)

# Supabase Auth integration (before_request guard, /login + /logout helpers)
import auth as _auth
_auth.init(app)

# ── SQLite Кирилл үсгийн case-insensitive хайлт ──────────────────────────────
from sqlalchemy import event as _sa_event
from sqlalchemy.pool import Pool as _Pool

@_sa_event.listens_for(_Pool, "connect")
def _register_unicode_lower(dbapi_conn, _):
    """SQLite-д Python unicode lower() функц бүртгэнэ — Кирилл үсэгт ажиллана"""
    dbapi_conn.create_function("uni_lower", 1, lambda x: x.lower() if x else "")

def _ilike(col, pattern):
    """Кирилл үсгийн case-insensitive LIKE (uni_lower ашиглана)"""
    from sqlalchemy import func as _f
    return _f.uni_lower(col).like(pattern.lower())

# ── Кирилл ↔ Латин транслитераци (бизнесийн нэр хайлтанд) ───────────────────
_C2L = {
    'А':'A','Б':'B','В':'V','Г':'G','Д':'D','Е':'E','Ё':'YO','Ж':'J',
    'З':'Z','И':'I','Й':'Y','К':'K','Л':'L','М':'M','Н':'N','О':'O',
    'П':'P','Р':'R','С':'S','Т':'T','У':'U','Ф':'F','Х':'H','Ц':'TS',
    'Ч':'CH','Ш':'SH','Щ':'SCH','Ъ':'','Ы':'Y','Ь':'','Э':'E','Ю':'YU',
    'Я':'YA','Ө':'O','Ү':'U',
}
_C2L.update({k.lower(): v.lower() for k, v in _C2L.items()})

_L2C = {
    'A':'А','B':'Б','C':'К','D':'Д','E':'Э','F':'Ф','G':'Г','H':'Х',
    'I':'И','J':'Ж','K':'К','L':'Л','M':'М','N':'Н','O':'О','P':'П',
    'Q':'К','R':'Р','S':'С','T':'Т','U':'У','V':'В','W':'В','X':'КС',
    'Y':'Й','Z':'З',
}
_L2C.update({k.lower(): v for k, v in _L2C.items()})

# Нэмэлт орлуулга: К↔C, К↔K, Х↔H зэрэг дуу хоолойн ижил үсгүүд
_PHONETIC_ALTS = [
    ('K', 'C'), ('K', 'Q'),
    ('V', 'W'),
    ('I', 'Y'),
    ('J', 'Y'),
    ('Z', 'S'),
    ('F', 'PH'),
]

def _translit(s):
    """Кирилл → Латин эсвэл Латин → Кирилл хувиргана"""
    if not s: return ''
    is_cyr = any(c in _C2L for c in s)
    if is_cyr:
        return ''.join(_C2L.get(c, c) for c in s)
    else:
        return ''.join(_L2C.get(c, c) for c in s)

def _search_variants(term):
    """Хайлтын нэрнээс бүх боломжит хувилбарыг үүсгэнэ (Кирилл+Латин+фонетик)"""
    t = term.strip()
    variants = {t}
    # Кирилл ↔ Латин хөрвүүлэлт
    alt = _translit(t)
    if alt: variants.add(alt)
    # Фонетик орлуулга (жишээ: K↔C)
    upper_t = t.upper()
    upper_alt = alt.upper() if alt else ''
    for a, b in _PHONETIC_ALTS:
        for base in [upper_t, upper_alt]:
            if a in base:
                variants.add(base.replace(a, b))
            if b in base:
                variants.add(base.replace(b, a))
    return variants

def _partner_search_conds(col, term, alias_col=None):
    """Кирилл+Латин+фонетик бүх вариантаар OR хайлт хийнэ.
    alias_col өгвөл aliases JSON баганыг ч хайна."""
    conds = []
    seen = set()
    for v in _search_variants(term):
        vl = v.lower()
        if vl not in seen:
            seen.add(vl)
            conds.append(_ilike(col, f'%{v}%'))
            # aliases JSON баганыг ч шалгана
            if alias_col is not None:
                from sqlalchemy import func as _f
                conds.append(_f.uni_lower(alias_col).like(f'%{v.lower()}%'))
    return db.or_(*conds)

@app.context_processor
def inject_now():
    company = {
        'name':         '"Түмэн Тээх" ХХК',
        'name_upper':   'ТҮМЭН ТЭЭХ ХХК',
        'name_en':      'TUEMEN TEEKH LLC',
        'type':         'Хязгаарлагдмал хариуцлагатай компани',
        'state_reg':    '000256787',
        'register':     '6906192',
        'tax_id':       '9011842351',
        'founded':      '2022.08.24',
        'director':     'Э.Өнөрсайхан',
        'director_last':'Энхтайван',
        'director_full':'Энхтайван овогтой Өнөрсайхан',
        'accountant':   'Т.Нацагдорж',
        'address':      'Монгол ТВ таур 1106 тоот, Сүхбаатар дүүрэг, 1 хороо, Чингисийн өргөн чөлөө, Улаанбаатар',
        'phone':        '77751111',
        'phone_director': '88881423',
        'phone_accountant': '99044424',
        'email':        'hello@tumentech.mn',
        'website':      'www.tumentech.mn',
        'capital':      '20,000,000',
        'bank_name':    'Худалдаа Хөгжилийн банк',
        'bank_account': '411096635',
        'bank_iban':    'MN320004000411096635',
    }
    lang = session.get('lang', 'mn')
    return {'now': datetime.now(), 'company': company, 'lang': lang, 't': get_t(lang)}


@app.route('/set-lang/<lang>')
def set_lang(lang):
    if lang in ('mn', 'en'):
        session['lang'] = lang
    return redirect(request.referrer or url_for('index'))


# ── Healthcheck (public) ────────────────────────────────────────────────────
@app.route('/healthz')
def healthz():
    return {'status': 'ok'}, 200


# ── Authentication: /login + /logout ────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login_view():
    next_url = request.values.get('next') or url_for('index')
    if request.method == 'POST':
        email    = (request.form.get('email') or '').strip().lower()
        password = request.form.get('password') or ''
        if not email or not password:
            return render_template('login.html', error='И-мэйл болон нууц үгээ оруулна уу',
                                   email=email, next_url=next_url)
        try:
            payload = _auth.sign_in(email, password)
        except _auth.AuthError as e:
            return render_template('login.html', error=str(e),
                                   email=email, next_url=next_url)

        # Resolve user's company + role
        access = payload.get('access_token')
        user   = payload.get('user') or {}
        company_id, role = _auth.fetch_user_company(user.get('id'), access)
        if not company_id:
            # No company assignment → block login
            return render_template('login.html',
                error='Энэ хэрэглэгчид компанийн эрх олгогдоогүй байна. Админд хандана уу.',
                email=email, next_url=next_url)

        _auth.store_session(payload, company_id=company_id, role=role)
        return redirect(next_url)

    return render_template('login.html', next_url=next_url)


@app.route('/logout', methods=['GET', 'POST'])
def logout_view():
    token = session.get('sb_access_token')
    _auth.sign_out(token)
    _auth.clear_session()
    session.pop('lang', None)
    return redirect(url_for('login_view'))


# ── Profile: view + change password ─────────────────────────────────────────
@app.route('/profile', methods=['GET'])
def profile_view():
    return render_template('profile.html', user=g.user)


@app.route('/profile/change-password', methods=['POST'])
def profile_change_password():
    current_password = request.form.get('current_password') or ''
    new_password     = request.form.get('new_password') or ''
    confirm_password = request.form.get('confirm_password') or ''

    if not current_password or not new_password:
        return render_template('profile.html', user=g.user,
                               error='Бүх талбарыг бөглөнө үү')
    if new_password != confirm_password:
        return render_template('profile.html', user=g.user,
                               error='Шинэ нууц үгүүд таарахгүй байна')
    if len(new_password) < 8:
        return render_template('profile.html', user=g.user,
                               error='Шинэ нууц үг 8+ тэмдэгт байх ёстой')
    if new_password == current_password:
        return render_template('profile.html', user=g.user,
                               error='Шинэ нууц үг хуучнаасаа ялгаатай байх ёстой')

    # Verify current password
    if not _auth.verify_password(g.user['email'], current_password):
        return render_template('profile.html', user=g.user,
                               error='Одоогийн нууц үг буруу байна')

    # Change password using user's own access token
    try:
        _auth.update_password(session.get('sb_access_token'), new_password)
    except _auth.AuthError as e:
        return render_template('profile.html', user=g.user, error=str(e))

    return render_template('profile.html', user=g.user,
                           success='Нууц үг амжилттай солигдлоо. Дараа удаа шинэ нууц үгээр нэвтэрнэ үү.')


# ══════════════════════════════════════════════════════════════════════════════
# МӨНГӨН УРСГАЛЫН КОДЛОХ ДҮРЭМ — Түмэн Тээх ХХК
# DT/KT холболтын автомат дүрмүүд
# ══════════════════════════════════════════════════════════════════════════════
def _bank_code(bank_str):
    """Банкны нэрнээс данс кодыг тодорхойлно"""
    if not bank_str: return '1101'
    b = str(bank_str)
    if '411096635' in b or 'ТДБ' in b or 'ХХБ' in b: return '1102'
    if '1175156757' in b or 'Голомт' in b:             return '1101'
    if '9006906192' in b or 'М банк' in b:             return '1103'
    return '1101'

# Орлогын кодлох дүрэм → DT: банк, KT: энд
CASHFLOW_INCOME_RULES = {
    '1.1.1': '6110',   # Үйлчилгээний орлого → Ачааны тээврийн орлого
    '1.1.2': '1210',   # Авлагын орлого      → Дансны авлага (барагдуулалт)
    '1.1.3': '8410',   # Хүүгийн орлого      → Хүүгийн орлого
    '1.1.4': '3310',   # Тээврийн буцаалт    → Дансны өглөг
    '5.1.1': '1211',   # Холбоотой байгуулагаас тооцоо
    '5.1.2': '4100',   # Зээлийн орлого      → Урт хугацааны зээл
    '5.1.3': '1240',   # Ажилтан зээлийн буцаалт
}

# Зарлагын кодлох дүрэм → DT: энд, KT: банк
CASHFLOW_EXPENSE_RULES = {
    '1.2.1':  '1250',   # Дотоод тээвр урьдчилгаа    → Бусад авлага (аванс)
    '1.2.2':  '3310',   # Өмнөх сарын тээвэр          → Дансны өглөг (AP)
    '2.1.1':  '3311',   # Цалингийн зардал            → Ажилтанд төлөх цалин
    '2.1.3':  '8214',   # Томилолт                    → Аялал, томилолтын зардал
    '2.1.5':  '8200',   # Сургалт                     → Ерөнхий зардал
    '2.1.10': '8210',   # Түрээсийн зардал
    '2.1.14': '8312',   # Банкны шимтгэл
    '2.2.1':  '3151',   # ААН орлогын татвар          → ААТ-ын өглөг
    '2.2.2':  '8660',   # Бусад зардал
    '2.2.3':  '3152',   # НӨАТ                        → НӨТ-ын өглөг
    '2.2.4':  '3321',   # НДШ / ЭМНДШ                 → НД-ын өглөг
    '3.2.1':  '1710',   # Программ хангамж            → Биет бус хөрөнгө
    '3.2.2':  '200501', # Тавилга, тоног төхөөрөмж    → Тавилга (ҮХ)
    '5.2.1':  '1211',   # Холбоотой байгуулагад тооцоо
    '5.2.2':  '4100',   # Зээлийн зарлага (эргэн төлөлт)
    '5.2.3':  '1240',   # Ажилтан зээл
}

def auto_link_cash_transaction(txn):
    """Нэг гүйлгээний DT/KT-г кодлох дүрмийн дагуу тодорхойлно"""
    bk = _bank_code(txn.bank)
    inc_cat = (txn.income_cat or '').strip()
    exp_cat = (txn.expense_cat or '').strip()
    inc = txn.income or 0
    exp = txn.expense or 0

    if inc > 0 and inc_cat in CASHFLOW_INCOME_RULES:
        return bk, CASHFLOW_INCOME_RULES[inc_cat]
    if exp > 0 and exp_cat in CASHFLOW_EXPENSE_RULES:
        return CASHFLOW_EXPENSE_RULES[exp_cat], bk
    return None, None

# ── Монгол тоо → үсгээр ─────────────────────────────────────────────────────
def num_to_words_mn(n):
    """Тоог Монгол үсгээр илэрхийлнэ. Жишээ: 50_000_000 → 'тавин сая'"""
    try:
        n = int(round(abs(float(n))))
    except (TypeError, ValueError):
        return ''
    if n == 0:
        return 'тэг'

    ONES_L = ['','нэг','хоёр','гурван','дөрвөн','таван','зургаан','долоон','найман','есөн']
    ONES_B = ['','нэг','хоёр','гурав','дөрөв','тав','зургаа','долоо','найм','ес']
    TENS_L = ['','арван','хорин','гучин','дөчин','тавин','жарин','далин','наян','ерэн']
    TENS_B = ['','арав','хорь','гуч','дөч','тавь','жар','дал','ная','ер']

    def chunk(num, before_unit):
        if num == 0:
            return ''
        h, r = divmod(num, 100)
        t, o = divmod(r, 10)
        parts = []
        if h:
            zu = 'зуун' if (r > 0 or before_unit) else 'зуу'
            parts.append(ONES_L[h] + ' ' + zu)
        if t and o:
            parts.append(TENS_L[t] + ' ' + ONES_L[o])
        elif t:
            parts.append(TENS_L[t] if before_unit else TENS_B[t])
        elif o:
            parts.append(ONES_L[o] if before_unit else ONES_B[o])
        return ' '.join(parts)

    result = []
    for div, unit in [(1_000_000_000, 'тэрбум'), (1_000_000, 'сая'), (1_000, 'мянга')]:
        q, n = divmod(n, div)
        if q:
            result.append(chunk(q, True) + ' ' + unit)
    if n:
        result.append(chunk(n, False))
    return ' '.join(result)

app.jinja_env.globals['num_to_words_mn'] = num_to_words_mn

# ===================== MODELS =====================

class Account(db.Model):
    __tablename__ = 'accounts'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(20), unique=True, nullable=False)
    name = db.Column(db.String(200), nullable=False)
    type = db.Column(db.String(50), nullable=False)  # asset, liability, equity, income, expense
    parent_id = db.Column(db.Integer, db.ForeignKey('accounts.id'), nullable=True)
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    children = db.relationship('Account', backref=db.backref('parent', remote_side=[id]))
    debit_lines = db.relationship('JournalLine', foreign_keys='JournalLine.debit_account_id', backref='debit_account')
    credit_lines = db.relationship('JournalLine', foreign_keys='JournalLine.credit_account_id', backref='credit_account')

class Journal(db.Model):
    __tablename__ = 'journals'
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    number = db.Column(db.String(50), unique=True)
    description = db.Column(db.String(500))
    reference = db.Column(db.String(200))
    status = db.Column(db.String(20), default='draft')  # draft, posted
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    lines = db.relationship('JournalLine', backref='journal', cascade='all, delete-orphan')

class JournalLine(db.Model):
    __tablename__ = 'journal_lines'
    id = db.Column(db.Integer, primary_key=True)
    journal_id = db.Column(db.Integer, db.ForeignKey('journals.id'), nullable=False)
    debit_account_id = db.Column(db.Integer, db.ForeignKey('accounts.id'), nullable=True)
    credit_account_id = db.Column(db.Integer, db.ForeignKey('accounts.id'), nullable=True)
    amount = db.Column(db.Float, nullable=False, default=0)
    description = db.Column(db.String(300))

# ---- Харилцагч (Partner) ----
class Partner(db.Model):
    __tablename__ = 'partners'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50))
    name = db.Column(db.String(200), nullable=False)
    register = db.Column(db.String(50))
    type = db.Column(db.String(20), default='both')  # customer, supplier, both
    phone = db.Column(db.String(50))
    email = db.Column(db.String(100))
    address = db.Column(db.String(300))
    is_active = db.Column(db.Boolean, default=True)
    aliases = db.Column(db.Text)   # JSON list of alternative names from master data
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    receivables = db.relationship('Receivable', backref='partner', lazy=True)
    payables = db.relationship('Payable', backref='partner', lazy=True)
    cash_txns = db.relationship('CashTransaction', backref='partner', lazy=True)

# ---- Авлага (Receivable) ----
class Receivable(db.Model):
    __tablename__ = 'receivables'
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    due_date = db.Column(db.Date)
    invoice_no = db.Column(db.String(100))
    partner_id = db.Column(db.Integer, db.ForeignKey('partners.id'), nullable=False)
    description = db.Column(db.String(300))
    amount = db.Column(db.Float, default=0)
    paid_amount = db.Column(db.Float, default=0)
    status = db.Column(db.String(20), default='open')  # open, partial, paid, cancelled
    responsible = db.Column(db.String(100))              # KAM / Хариуцагч нэр
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    @property
    def remaining(self):
        return self.amount - self.paid_amount

# ---- Өглөг (Payable) ----
class Payable(db.Model):
    __tablename__ = 'payables'
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    due_date = db.Column(db.Date)
    invoice_no = db.Column(db.String(100))
    partner_id = db.Column(db.Integer, db.ForeignKey('partners.id'), nullable=False)
    description = db.Column(db.String(300))
    amount = db.Column(db.Float, default=0)
    paid_amount = db.Column(db.Float, default=0)
    status = db.Column(db.String(20), default='open')  # open, partial, paid, cancelled
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    @property
    def remaining(self):
        return self.amount - self.paid_amount

# ---- Мөнгөн хөрөнгийн хөдөлгөөн (CashTransaction) ----
class CashTransaction(db.Model):
    __tablename__ = 'cash_transactions'
    id = db.Column(db.Integer, primary_key=True)
    row_num = db.Column(db.Integer)
    month = db.Column(db.Integer)
    year = db.Column(db.Integer)
    company = db.Column(db.String(100))
    date = db.Column(db.DateTime)
    bank = db.Column(db.String(100))
    description = db.Column(db.String(500))
    partner_name = db.Column(db.String(200))
    partner_acc = db.Column(db.String(100))
    exchange_rate = db.Column(db.Float, default=1)
    income = db.Column(db.Float, default=0)
    expense = db.Column(db.Float, default=0)
    income_cat = db.Column(db.String(50))
    expense_cat = db.Column(db.String(50))
    crm_code = db.Column(db.String(50))
    master_name = db.Column(db.String(200))
    debit_acc = db.Column(db.String(20))
    credit_acc = db.Column(db.String(20))
    partner_id = db.Column(db.Integer, db.ForeignKey('partners.id'), nullable=True)
    journal_id = db.Column(db.Integer, db.ForeignKey('journals.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ---- Үндсэн хөрөнгө (FixedAsset) ----
class FixedAsset(db.Model):
    __tablename__ = 'fixed_assets'
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(50), unique=True)
    name = db.Column(db.String(200), nullable=False)
    category = db.Column(db.String(100), default='Тоног төхөөрөмж')
    purchase_date = db.Column(db.Date)
    purchase_amount = db.Column(db.Float, default=0)
    useful_life = db.Column(db.Integer, default=5)        # жил
    salvage_value = db.Column(db.Float, default=0)
    accumulated_depreciation = db.Column(db.Float, default=0)
    location = db.Column(db.String(200))
    responsible_person = db.Column(db.String(100))
    status = db.Column(db.String(20), default='active')   # active, disposed, sold
    notes = db.Column(db.String(500))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    @property
    def book_value(self):
        return max(0, self.purchase_amount - self.accumulated_depreciation)

    @property
    def annual_depreciation(self):
        if self.useful_life and self.useful_life > 0:
            return (self.purchase_amount - self.salvage_value) / self.useful_life
        return 0

    @property
    def monthly_depreciation(self):
        return self.annual_depreciation / 12

    @property
    def depreciation_pct(self):
        if self.purchase_amount > 0:
            return min(100, self.accumulated_depreciation / self.purchase_amount * 100)
        return 0

class VatRecord(db.Model):
    __tablename__ = 'vat_records'
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    type = db.Column(db.String(20))   # 'out'=борлуулалт(цуглуулсан), 'in'=худалдан авалт(суутгал)
    ddtd = db.Column(db.String(50))   # ДДТД — eBarimt дугаар
    invoice_no = db.Column(db.String(100))
    partner_name = db.Column(db.String(200))
    partner_register = db.Column(db.String(50))
    amount = db.Column(db.Float, default=0)      # НӨАТ-гүй дүн
    vat_amount = db.Column(db.Float, default=0)  # НӨАТ дүн
    total_amount = db.Column(db.Float, default=0)
    paid_amount = db.Column(db.Float, default=0)
    remaining = db.Column(db.Float, default=0)
    tax_type = db.Column(db.String(50))           # Энгийн / Чөлөөлөгдөх
    source = db.Column(db.String(50))             # ИБАРИМТ / ПОС / Гар
    ebarimt_status = db.Column(db.String(50))     # Илгээгдсэн / Хүлээгдэж байна
    ebarimt_sent = db.Column(db.Boolean, default=False)
    month = db.Column(db.Integer)
    parent_ddtd   = db.Column(db.String(100), nullable=True)  # Толгой нэхэмжлэхийн ДДТД (хаалтын баримтад бичигдэнэ)
    partner_id    = db.Column(db.Integer, db.ForeignKey('partners.id'),    nullable=True)
    receivable_id = db.Column(db.Integer, db.ForeignKey('receivables.id'), nullable=True)
    partner    = db.relationship('Partner',    foreign_keys=[partner_id],    lazy=True)
    receivable = db.relationship('Receivable', foreign_keys=[receivable_id], lazy=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

# ---- Тулгалт (Reconciliation) ----
class Reconciliation(db.Model):
    __tablename__ = 'reconciliations'
    id            = db.Column(db.Integer, primary_key=True)
    partner_id    = db.Column(db.Integer, db.ForeignKey('partners.id'), nullable=False)
    vat_record_id = db.Column(db.Integer, db.ForeignKey('vat_records.id'),      nullable=True)
    cash_txn_id   = db.Column(db.Integer, db.ForeignKey('cash_transactions.id'), nullable=True)
    matched_amount= db.Column(db.Float, default=0)   # тулгасан дүн
    diff          = db.Column(db.Float, default=0)   # зөрүү (+ = төлөөгүй, - = илүү төлсөн)
    notes         = db.Column(db.String(300))
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    partner    = db.relationship('Partner',         foreign_keys=[partner_id])
    vat_record = db.relationship('VatRecord',       foreign_keys=[vat_record_id])
    cash_txn   = db.relationship('CashTransaction', foreign_keys=[cash_txn_id])

# ---- Нэхэмжлэл ↔ Банкны гүйлгээ тулгалт (InvoicePayment) ----
class InvoicePayment(db.Model):
    __tablename__ = 'invoice_payments'
    id            = db.Column(db.Integer, primary_key=True)
    receivable_id = db.Column(db.Integer, db.ForeignKey('receivables.id'),        nullable=False)
    cash_txn_id   = db.Column(db.Integer, db.ForeignKey('cash_transactions.id'),  nullable=False)
    amount        = db.Column(db.Float, default=0)
    notes         = db.Column(db.String(300))
    created_at    = db.Column(db.DateTime, default=datetime.utcnow)
    receivable    = db.relationship('Receivable',      foreign_keys=[receivable_id])
    cash_txn      = db.relationship('CashTransaction', foreign_keys=[cash_txn_id])

# ---- Ажилтан (Employee) ----
class Employee(db.Model):
    __tablename__ = 'employees'
    id              = db.Column(db.Integer, primary_key=True)
    name            = db.Column(db.String(60), nullable=False)   # Нэр
    lastname        = db.Column(db.String(60))                   # Овог
    tin             = db.Column(db.String(30))                   # РД
    company         = db.Column(db.String(100), default='ТҮМЭН ТЭЭХ ХХК')
    title           = db.Column(db.String(100))                  # Албан тушаал
    base_salary     = db.Column(db.Float, default=0)             # Үндсэн цалин
    adv_base        = db.Column(db.Float)                        # Урьдчилгааны суурь (None → base*0.4)
    phone           = db.Column(db.String(30))                   # Утасны дугаар
    phone_allowance = db.Column(db.Float, default=0)             # Утасны нэмэгдэл
    hire_date       = db.Column(db.String(20))                   # Ажилд орсон огноо
    exp_years       = db.Column(db.Integer, default=0)           # Ажлын жил
    bank            = db.Column(db.String(50))
    account         = db.Column(db.String(60))
    email           = db.Column(db.String(100))
    color           = db.Column(db.String(20), default='#3B82F6')
    is_active       = db.Column(db.Boolean, default=True)
    partner_id      = db.Column(db.Integer, db.ForeignKey('partners.id'), nullable=True)
    partner         = db.relationship('Partner', foreign_keys=[partner_id])
    salary_records  = db.relationship('SalaryRecord', backref='employee', lazy=True,
                                      order_by='SalaryRecord.year.desc(), SalaryRecord.month.desc()')

    @property
    def full_name(self):
        return f'{self.lastname[0]}. {self.name}' if self.lastname else self.name

# ---- Цалингийн бүртгэл (SalaryRecord) ----
class SalaryRecord(db.Model):
    __tablename__ = 'salary_records'
    id             = db.Column(db.Integer, primary_key=True)
    employee_id    = db.Column(db.Integer, db.ForeignKey('employees.id'), nullable=False)
    year           = db.Column(db.Integer, nullable=False)
    month          = db.Column(db.Integer, nullable=False)
    # Орц
    base_salary    = db.Column(db.Float, default=0)
    worked_hrs     = db.Column(db.Float, default=0)
    total_hrs      = db.Column(db.Float, default=176)
    phone          = db.Column(db.Float, default=0)
    sales_bonus    = db.Column(db.Float, default=0)    # Борлуулалтын нэмэгдэл
    leave_pay      = db.Column(db.Float, default=0)    # ЭА нэмэгдэл
    other_ded      = db.Column(db.Float, default=0)    # Бусад суутгал
    adv_override   = db.Column(db.Float)               # Урьдчилгаа (None → base*0.4)
    # Тооцоолол
    bod_salary     = db.Column(db.Float, default=0)    # Бодогдсон цалин
    total_income   = db.Column(db.Float, default=0)    # Нийт орлого
    emndsh         = db.Column(db.Float, default=0)    # ЭМНДШ+НД 11.5%
    hhoat_ded      = db.Column(db.Float, default=0)    # 23.1 хөнгөлөлт
    hhoat          = db.Column(db.Float, default=0)    # ХХОАТ
    advance        = db.Column(db.Float, default=0)    # Урьдчилгаа
    net_pay        = db.Column(db.Float, default=0)    # Гарт авах
    status         = db.Column(db.String(20), default='draft')  # draft/paid
    notes          = db.Column(db.String(300))
    created_at     = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at     = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

# ─── Цалин тооцооны тогтмол ───────────────────────────────────────────────────
MONTH_HOURS = {1:136,2:152,3:168,4:176,5:160,6:168,7:184,8:168,9:176,10:184,11:160,12:184}
EMNDSH_CAP  = 7_920_000

def art231(inc):
    """ХХОАТ-ын 23.1-р зүйлийн хөнгөлөлт (сарын орлогоор)"""
    if inc <= 500_000:   return 20_000
    if inc <= 1_000_000: return 18_000
    if inc <= 1_500_000: return 16_000
    if inc <= 2_000_000: return 14_000
    if inc <= 2_500_000: return 12_000
    if inc <= 3_000_000: return 10_000
    return 0

EMNDSH_ORG_RATE = 0.125   # Байгууллагын ЭМНДШ + НД: 12.5%

def calc_salary(base, worked_hrs, month, phone=0, sales=0, leave=0, adv_override=None):
    """Цалин тооцоолно — цаг_бүртгэл.html-ийн calcSalary()-тай яг ижил логик"""
    total_hrs  = MONTH_HOURS.get(month, 176)
    bod        = (base / total_hrs * worked_hrs) if total_hrs > 0 else 0
    niit       = bod + phone + sales + leave
    emndsh     = min(niit, EMNDSH_CAP) * 0.115          # ажилтны хувь
    emndsh_org = min(niit, EMNDSH_CAP) * EMNDSH_ORG_RATE  # байгууллагын хувь
    ded23      = art231(niit)
    hhoat      = max(0.0, (niit - emndsh) * 0.1 - ded23)
    adv        = adv_override if adv_override is not None else base * 0.4
    gart       = niit - emndsh - hhoat - adv
    return {
        'bod':        round(bod),
        'niit':       round(niit),
        'emndsh':     round(emndsh),
        'emndsh_org': round(emndsh_org),
        'ded23':      round(ded23),
        'hhoat':      round(hhoat),
        'adv':        round(adv),
        'gart':       round(gart),
        'total_hrs':  total_hrs,
    }

def seed_employees():
    """цаг_бүртгэл.html-ийн DEFAULT_EMPLOYEES-аас ТҮМЭН ТЭЭХ ажилтнуудыг нэмнэ"""
    if Employee.query.filter_by(company='ТҮМЭН ТЭЭХ ХХК').count() > 0:
        return  # Аль хэдийн нэмэгдсэн
    tt_emps = [
        dict(name='Отгонбаатар', lastname='Тогтсүх',     tin='112293420363', base_salary=5400000, phone_allowance=40000,  hire_date='2021-09-30', exp_years=13, adv_base=None,    title='Үйл ажиллагааны захирал',  bank='Хас',    account='MN590032005000866447', color='#EF4444'),
        dict(name='Уранзаяа',    lastname='Жаргалсайхан', tin='111673170867', base_salary=2900000, phone_allowance=200000, hire_date='2025-04-03', exp_years=1,  adv_base=None,    title='ХАБ-н менежер',            bank='Голомт', account='MN310015002905160782', color='#F59E0B'),
        dict(name='Нямдорж',     lastname='Мягмардорж',  tin='110733681978', base_salary=3500000, phone_allowance=40000,  hire_date='2022-02-01', exp_years=15, adv_base=3000000, title='Харилцагчийн менежер',      bank='Голомт', account='MN370015002135100372', color='#06B6D4'),
        dict(name='Одонтунгалаг',lastname='Даваадорж',   tin='112584431122', base_salary=3000000, phone_allowance=0,      hire_date='2025-08-04', exp_years=9,  adv_base=2800000, title='Тээврийн менежер',          bank='Хаан',   account='MN310005005111068934', color='#84CC16'),
        dict(name='Тэргэл',      lastname='Ганзам',       tin='111038832317', base_salary=2800000, phone_allowance=0,      hire_date='2024-01-01', exp_years=5,  adv_base=None,    title='Transportation manager',    bank='Голомт', account='040015002035100174',  color='#EC4899'),
        dict(name='Баяраа',      lastname='Оюундулам',   tin='620198632328', base_salary=3200000, phone_allowance=40000,  hire_date='2023-09-04', exp_years=7,  adv_base=2800000, title='Харилцагчийн менежер',      bank='Голомт', account='MN020015001105307098', color='#6366F1'),
    ]
    for d in tt_emps:
        e = Employee(company='ТҮМЭН ТЭЭХ ХХК', email='', **d)
        # Partner-тай холбох (TIN-аар)
        p = Partner.query.filter_by(is_active=True).filter(
            db.func.upper(Partner.name).like(f'%{d["name"].upper()}%')
        ).first()
        if p:
            e.partner_id = p.id
        db.session.add(e)
    db.session.commit()

# ===================== HELPERS =====================

def get_next_journal_number():
    last = Journal.query.order_by(Journal.id.desc()).first()
    year = datetime.now().year
    if last:
        return f"JNL-{year}-{last.id + 1:04d}"
    return f"JNL-{year}-0001"

def get_account_balance(account_id):
    debit_sum = db.session.query(db.func.sum(JournalLine.amount)).filter(
        JournalLine.debit_account_id == account_id,
        JournalLine.journal.has(status='posted')
    ).scalar() or 0
    credit_sum = db.session.query(db.func.sum(JournalLine.amount)).filter(
        JournalLine.credit_account_id == account_id,
        JournalLine.journal.has(status='posted')
    ).scalar() or 0
    return debit_sum - credit_sum

def get_dashboard_stats():
    from sqlalchemy import func

    # Journals
    total_accounts = Account.query.filter_by(is_active=True).count()
    total_journals = Journal.query.count()
    posted_journals = Journal.query.filter_by(status='posted').count()
    draft_journals = Journal.query.filter_by(status='draft').count()

    income_accounts = Account.query.filter_by(type='income').all()
    expense_accounts = Account.query.filter_by(type='expense').all()
    total_income = sum(abs(get_account_balance(a.id)) for a in income_accounts)
    total_expense = sum(abs(get_account_balance(a.id)) for a in expense_accounts)

    # Cash (bank) stats
    cash_income = db.session.query(func.sum(CashTransaction.income)).scalar() or 0
    cash_expense = db.session.query(func.sum(CashTransaction.expense)).scalar() or 0
    cash_count = CashTransaction.query.count()

    # Monthly cash breakdown
    monthly_cash = db.session.query(
        CashTransaction.month,
        func.sum(CashTransaction.income).label('income'),
        func.sum(CashTransaction.expense).label('expense')
    ).group_by(CashTransaction.month).order_by(CashTransaction.month).all()

    # Авлага / Өглөг — дансны үлдэгдлээс тооцно (Receivable хүснэгт хоосон байвал)
    manual_recv = db.session.query(func.sum(Receivable.amount)).scalar() or 0
    if manual_recv > 0:
        recv_paid = db.session.query(func.sum(Receivable.paid_amount)).scalar() or 0
        recv_total    = manual_recv
        recv_remaining= manual_recv - recv_paid
        recv_open     = Receivable.query.filter(Receivable.status != 'paid').count()
    else:
        # 12xx дансдын үлдэгдлийг нэгтгэнэ (Авлага)
        recv_accs = Account.query.filter(
            Account.code.like('12%'), Account.is_active == True
        ).all()
        recv_total     = sum(max(0, get_account_balance(a.id)) for a in recv_accs)
        recv_remaining = recv_total
        recv_open      = len([a for a in recv_accs if get_account_balance(a.id) > 1])

    manual_pay = db.session.query(func.sum(Payable.amount)).scalar() or 0
    if manual_pay > 0:
        pay_paid = db.session.query(func.sum(Payable.paid_amount)).scalar() or 0
        pay_total     = manual_pay
        pay_remaining = manual_pay - pay_paid
        pay_open      = Payable.query.filter(Payable.status != 'paid').count()
    else:
        # 31xx-34xx дансдын үлдэгдлийг нэгтгэнэ (Өглөг — KT тал → сөрөг үлдэгдэл)
        pay_accs = Account.query.filter(
            Account.is_active == True
        ).filter(
            db.or_(Account.code.like('31%'), Account.code.like('32%'),
                   Account.code.like('33%'), Account.code.like('34%'))
        ).all()
        pay_total     = sum(max(0, -get_account_balance(a.id)) for a in pay_accs)
        pay_remaining = pay_total
        pay_open      = len([a for a in pay_accs if get_account_balance(a.id) < -1])

    # VAT stats
    out_vat = db.session.query(func.sum(VatRecord.vat_amount)).filter_by(type='out').scalar() or 0
    in_vat  = db.session.query(func.sum(VatRecord.vat_amount)).filter_by(type='in').scalar() or 0
    out_cnt = VatRecord.query.filter_by(type='out').count()
    in_cnt  = VatRecord.query.filter_by(type='in').count()

    # Partners
    total_partners = Partner.query.filter_by(is_active=True).count()

    # Top partners by income (from cash)
    top_partners = db.session.query(
        CashTransaction.partner_name,
        func.sum(CashTransaction.income).label('income'),
        func.sum(CashTransaction.expense).label('expense')
    ).filter(CashTransaction.partner_name != None, CashTransaction.partner_name != '')\
     .group_by(CashTransaction.partner_name)\
     .order_by(func.sum(CashTransaction.income).desc()).limit(5).all()

    # Bank breakdown (raw, for legacy use)
    by_bank = db.session.query(
        CashTransaction.bank,
        func.sum(CashTransaction.income).label('income'),
        func.sum(CashTransaction.expense).label('expense')
    ).group_by(CashTransaction.bank).order_by(CashTransaction.bank).all()

    # ── Банкны үлдэгдэл нэгтгэл ─────────────────────────────────────────────
    from sqlalchemy import text as _text
    _open_rows = db.session.execute(_text(
        "SELECT a.code, "
        "SUM(CASE WHEN jl.debit_account_id = a.id  THEN jl.amount ELSE 0 END) - "
        "SUM(CASE WHEN jl.credit_account_id = a.id THEN jl.amount ELSE 0 END) AS balance "
        "FROM accounts a "
        "JOIN journal_lines jl ON (jl.debit_account_id = a.id OR jl.credit_account_id = a.id) "
        "JOIN journals j ON j.id = jl.journal_id "
        "WHERE a.code IN ('1101','1102','1103') "
        "AND j.status = 'posted' AND j.date <= '2025-12-31' "
        "GROUP BY a.code"
    )).fetchall()
    _open_map = {r[0]: float(r[1] or 0) for r in _open_rows}

    def _bcode(b):
        if not b: return '1101'
        if '411096635' in b or 'ТДБ' in b or 'ХХБ' in b: return '1102'
        if '1175156757' in b or 'Голомт' in b:             return '1101'
        if '9006906192' in b or 'М банк' in b:             return '1103'
        return 'other'

    _inc_by_code = {}; _exp_by_code = {}
    for b in by_bank:
        c = _bcode(b.bank)
        _inc_by_code[c] = _inc_by_code.get(c, 0) + float(b.income or 0)
        _exp_by_code[c] = _exp_by_code.get(c, 0) + float(b.expense or 0)

    _BANK_DEFS = [('1101','Голомт банк'),('1102','ХХБ / ТДБ'),('1103','М банк')]
    bank_balances = []
    for code, label in _BANK_DEFS:
        op  = _open_map.get(code, 0)
        inc = _inc_by_code.get(code, 0)
        exp = _exp_by_code.get(code, 0)
        bank_balances.append({
            'code': code, 'label': label,
            'open': op, 'income': inc, 'expense': exp,
            'close': op + inc - exp,
        })
    _total_open  = sum(b['open']  for b in bank_balances)
    _total_inc   = sum(b['income']  for b in bank_balances)
    _total_exp   = sum(b['expense'] for b in bank_balances)
    bank_totals  = {'open': _total_open, 'income': _total_inc,
                    'expense': _total_exp, 'close': _total_open + _total_inc - _total_exp}

    return {
        'total_accounts': total_accounts,
        'total_journals': total_journals,
        'posted_journals': posted_journals,
        'draft_journals': draft_journals,
        'total_income': total_income,
        'total_expense': total_expense,
        'net_profit': total_income - total_expense,
        # cash
        'cash_income': cash_income,
        'cash_expense': cash_expense,
        'cash_net': cash_income - cash_expense,
        'cash_count': cash_count,
        'monthly_cash': monthly_cash,
        'by_bank': by_bank,
        'bank_balances': bank_balances,
        'bank_totals': bank_totals,
        'top_partners': top_partners,
        # ar/ap
        'recv_total': recv_total,
        'recv_remaining': recv_remaining,
        'recv_open': recv_open,
        'pay_total': pay_total,
        'pay_remaining': pay_remaining,
        'pay_open': pay_open,
        'total_partners': total_partners,
        # vat
        'out_vat': out_vat,
        'in_vat': in_vat,
        'net_vat': out_vat - in_vat,
        'out_cnt': out_cnt,
        'in_cnt': in_cnt,
    }

# ===================== ROUTES =====================

@app.route('/')
def index():
    stats = get_dashboard_stats()
    recent_journals = Journal.query.order_by(Journal.created_at.desc()).limit(10).all()
    return render_template('index.html', stats=stats, recent_journals=recent_journals)

# --- Accounts ---
@app.route('/accounts')
def accounts():
    accounts = Account.query.filter_by(is_active=True).order_by(Account.code).all()
    return render_template('accounts.html', accounts=accounts)

@app.route('/accounts/add', methods=['GET', 'POST'])
def add_account():
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        name = request.form.get('name', '').strip()
        type_ = request.form.get('type', '')
        parent_id = request.form.get('parent_id') or None
        if not code or not name or not type_:
            flash('Бүх талбарыг бөглөнө үү!', 'danger')
            return redirect(url_for('add_account'))
        if Account.query.filter_by(code=code).first():
            flash('Энэ кодтой данс аль хэдийн байна!', 'danger')
            return redirect(url_for('add_account'))
        acc = Account(code=code, name=name, type=type_, parent_id=parent_id)
        db.session.add(acc)
        db.session.commit()
        flash(f'"{name}" данс амжилттай нэмэгдлээ!', 'success')
        return redirect(url_for('accounts'))
    parent_accounts = Account.query.filter_by(is_active=True).order_by(Account.code).all()
    return render_template('account_form.html', account=None, parent_accounts=parent_accounts)

@app.route('/accounts/edit/<int:id>', methods=['GET', 'POST'])
def edit_account(id):
    acc = Account.query.get_or_404(id)
    if request.method == 'POST':
        acc.code = request.form.get('code', acc.code).strip()
        acc.name = request.form.get('name', acc.name).strip()
        acc.type = request.form.get('type', acc.type)
        acc.parent_id = request.form.get('parent_id') or None
        db.session.commit()
        flash('Данс амжилттай засагдлаа!', 'success')
        return redirect(url_for('accounts'))
    parent_accounts = Account.query.filter(Account.id != id, Account.is_active == True).order_by(Account.code).all()
    return render_template('account_form.html', account=acc, parent_accounts=parent_accounts)

@app.route('/accounts/delete/<int:id>', methods=['POST'])
def delete_account(id):
    acc = Account.query.get_or_404(id)
    acc.is_active = False
    db.session.commit()
    flash('Данс устгагдлаа!', 'success')
    return redirect(url_for('accounts'))

@app.route('/api/accounts')
def api_accounts():
    accounts = Account.query.filter_by(is_active=True).order_by(Account.code).all()
    return jsonify([{'id': a.id, 'code': a.code, 'name': a.name, 'type': a.type} for a in accounts])

# --- Journals ---
@app.route('/journals')
def journals():
    journals = Journal.query.order_by(Journal.date.desc(), Journal.id.desc()).all()
    return render_template('journals.html', journals=journals)

@app.route('/journals/add', methods=['GET', 'POST'])
def add_journal():
    if request.method == 'POST':
        date_str = request.form.get('date')
        description = request.form.get('description', '')
        reference = request.form.get('reference', '')
        lines_json = request.form.get('lines_data', '[]')
        lines = json.loads(lines_json)

        if not date_str or not lines:
            flash('Огноо болон гүйлгээний мөрүүдийг оруулна уу!', 'danger')
            return redirect(url_for('add_journal'))

        journal = Journal(
            date=datetime.strptime(date_str, '%Y-%m-%d').date(),
            number=get_next_journal_number(),
            description=description,
            reference=reference,
            status='draft'
        )
        db.session.add(journal)
        db.session.flush()

        total_debit = 0
        total_credit = 0
        for line in lines:
            amount = float(line.get('amount', 0))
            debit_id = line.get('debit_account_id') or None
            credit_id = line.get('credit_account_id') or None
            if debit_id:
                total_debit += amount
            if credit_id:
                total_credit += amount
            jl = JournalLine(
                journal_id=journal.id,
                debit_account_id=debit_id,
                credit_account_id=credit_id,
                amount=amount,
                description=line.get('description', '')
            )
            db.session.add(jl)

        db.session.commit()
        flash(f'Гүйлгээ {journal.number} амжилттай хадгалагдлаа!', 'success')
        return redirect(url_for('journals'))

    accounts = Account.query.filter_by(is_active=True).order_by(Account.code).all()
    return render_template('journal_form.html', journal=None, accounts=accounts, today=date.today().isoformat())

@app.route('/journals/<int:id>')
def view_journal(id):
    journal = Journal.query.get_or_404(id)
    return render_template('journal_view.html', journal=journal)

@app.route('/journals/<int:id>/post', methods=['POST'])
def post_journal(id):
    journal = Journal.query.get_or_404(id)
    journal.status = 'posted'
    db.session.commit()
    flash(f'{journal.number} гүйлгээ баталгаажлаа!', 'success')
    return redirect(url_for('view_journal', id=id))

@app.route('/journals/<int:id>/delete', methods=['POST'])
def delete_journal(id):
    journal = Journal.query.get_or_404(id)
    if journal.status == 'posted':
        flash('Баталгаажсан гүйлгээг устгах боломжгүй!', 'danger')
        return redirect(url_for('journals'))
    db.session.delete(journal)
    db.session.commit()
    flash('Гүйлгээ устгагдлаа!', 'success')
    return redirect(url_for('journals'))

# ── helpers for financial statements ──
def acct_bal(code):
    """Get balance for account by code"""
    a = Account.query.filter_by(code=code, is_active=True).first()
    if not a: return 0
    return get_account_balance(a.id)

def group_bal(codes):
    """Sum balances of multiple account codes"""
    return sum(acct_bal(c) for c in codes)

def subtree_bal(root_code):
    """Sum all accounts starting with root_code prefix"""
    accts = Account.query.filter(
        Account.code.like(root_code + '%'), Account.is_active == True
    ).all()
    return sum(get_account_balance(a.id) for a in accts)

# --- Reports ---
@app.route('/reports')
def reports():
    return render_template('reports.html')

@app.route('/reports/balance-sheet')
def balance_sheet():
    from sqlalchemy import func as sqlfunc

    # ── Огноо параметр (default: 2025-12-31 = OB огноо) ─────────────────────
    date_str = request.args.get('as_of', '2025-12-31')
    try:
        report_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except Exception:
        report_date = date(2025, 12, 31)

    # ── Огноогоор шүүсэн баланс тооцох ─────────────────────────────────────
    def _bal(account_id, asof):
        j_ids = db.session.query(Journal.id).filter(
            Journal.status == 'posted',
            Journal.date <= asof
        ).subquery()
        dt = db.session.query(sqlfunc.sum(JournalLine.amount)).filter(
            JournalLine.debit_account_id == account_id,
            JournalLine.journal_id.in_(j_ids)
        ).scalar() or 0
        kt = db.session.query(sqlfunc.sum(JournalLine.amount)).filter(
            JournalLine.credit_account_id == account_id,
            JournalLine.journal_id.in_(j_ids)
        ).scalar() or 0
        return dt - kt

    def ab(code, asof):
        a = Account.query.filter_by(code=code, is_active=True).first()
        return _bal(a.id, asof) if a else 0

    def gb(codes, asof):
        return sum(ab(c, asof) for c in codes)

    def stb(root, asof):
        accts = Account.query.filter(
            Account.code.like(root + '%'), Account.is_active == True
        ).all()
        return sum(_bal(a.id, asof) for a in accts)

    def compute_bs(asof):
        # ── Эргэлтийн хөрөнгө ───────────────────────────────────────────────
        cash       = gb(['1110','1101','1102','1103'], asof)
        recv_trade = ab('1210', asof) + ab('1211', asof)
        recv_tax   = stb('135', asof)           # 1350, 1351 (урьдчилан ААТ), 1352 (НӨАТ авлага)
        recv_other = ab('1250', asof) + ab('1240', asof)
        prepaid    = ab('1260', asof)
        inventory  = gb(['1310','1320','1330'], asof)
        total_current = cash + recv_trade + recv_tax + recv_other + prepaid + inventory

        # ── Эргэлтийн бус хөрөнгө ───────────────────────────────────────────
        # 200xxx = ҮХ өртөг, 201xxx = ҮХ хуримтлагдсан элэгдэл (credit)
        fixed_assets = stb('200', asof) + stb('201', asof) + stb('16', asof)
        intangible   = ab('1710', asof) + ab('1510', asof)
        total_noncurrent = fixed_assets + intangible
        total_assets = total_current + total_noncurrent

        # ── Богино хугацаат өр ──────────────────────────────────────────────
        ap_trade    = -ab('3310', asof)
        payroll_pay = -(ab('3311', asof) + ab('3313', asof))
        tax_pay     = -(ab('3100', asof) + ab('3141', asof) + ab('3151', asof) + ab('3152', asof))
        ndsh_pay    = -(ab('3321', asof) + ab('3322', asof))
        other_short = -(ab('3210', asof) + ab('3330', asof) + ab('3333', asof) +
                        ab('3332', asof) + ab('3410', asof))
        total_short = ap_trade + payroll_pay + tax_pay + ndsh_pay + other_short

        # ── Урт хугацаат өр ─────────────────────────────────────────────────
        lt_loans    = -(ab('4100', asof) + ab('4110', asof))
        lt_other    = -(ab('4210', asof) + ab('4211', asof) + ab('4310', asof) + ab('4320', asof))
        total_long  = lt_loans + lt_other
        total_liab  = total_short + total_long

        # ── Эздийн өмч ──────────────────────────────────────────────────────
        capital     = -ab('5110', asof)
        add_capital = 0
        retained    = -(ab('5200', asof) + ab('5210', asof))

        # ── Тайлант үеийн цэвэр ашиг/алдагдал ──────────────────────────────
        # 6xxx орлого, 7xxx COGS, 8xxx G&A, 9xxx татварын зардал
        # Хаалтын бичилт хийгдээгүй тул P&L дансуудын үлдэгдлийг
        # өмчийн хэсэгт "тайлант үеийн ашиг" болгон нэмнэ
        current_profit = -(stb('6', asof) + stb('7', asof) + stb('8', asof) + stb('9', asof))

        total_equity      = capital + add_capital + retained + current_profit
        total_liab_equity = total_liab + total_equity

        return dict(
            cash=cash, recv_trade=recv_trade, recv_tax=recv_tax, recv_other=recv_other,
            prepaid=prepaid, inventory=inventory, total_current=total_current,
            fixed_assets=fixed_assets, intangible=intangible, total_noncurrent=total_noncurrent,
            total_assets=total_assets,
            ap_trade=ap_trade, payroll_pay=payroll_pay, tax_pay=tax_pay, ndsh_pay=ndsh_pay,
            other_short=other_short, total_short=total_short,
            lt_loans=lt_loans, lt_other=lt_other, total_long=total_long, total_liab=total_liab,
            capital=capital, add_capital=add_capital, retained=retained,
            current_profit=current_profit,
            total_equity=total_equity, total_liab_equity=total_liab_equity
        )

    bs = compute_bs(report_date)

    # ── 2025-12-31 харьцуулалт (Excel-ийн баримтаас) ───────────────────────
    bs25 = dict(
        cash=173733062.20, recv_trade=171629688.95, recv_tax=0,
        recv_other=602156327.00, prepaid=197719079.71, inventory=0,
        total_current=1145238157.86,
        fixed_assets=28685851.00, intangible=0, total_noncurrent=28685851.00,
        total_assets=1173924008.86,
        ap_trade=391771972.26, payroll_pay=28406726.42, tax_pay=238952813.31,
        ndsh_pay=7211944.25, other_short=0, total_short=651919568.00,
        lt_loans=0, lt_other=0, total_long=0, total_liab=651919568.00,
        capital=20000, add_capital=0, retained=521984441.13,
        total_equity=522004441.13, total_liab_equity=1173924009.13
    )
    return render_template('balance_sheet.html', bs=bs, bs25=bs25,
                           report_date=report_date, prev_label='2025-12-31')

@app.route('/reports/cit')
def cit_report():
    """Аж ахуйн нэгжийн орлогын албан татварын тайлан (ААНОАТ)"""
    from sqlalchemy import func as _sf

    # ── Огноо ────────────────────────────────────────────────────────────────
    cur_year = date.today().year
    year_str  = request.args.get('year', str(cur_year))
    q_str     = request.args.get('quarter', '')   # '' = жилийн, '1'..'4' = улирлын
    try:
        year = int(year_str)
    except Exception:
        year = cur_year

    if q_str in ('1','2','3','4'):
        q = int(q_str)
        month_start = (q - 1) * 3 + 1
        month_end   = q * 3
        df = date(year, month_start, 1)
        import calendar
        dt = date(year, month_end, calendar.monthrange(year, month_end)[1])
        period_label = f'{year} оны {q}-р улирал'
    else:
        q_str = ''
        df = date(year, 1, 1)
        dt = date(year, 12, 31)
        period_label = f'{year} оны жилийн тайлан'

    # ── Дансны үлдэгдэл тооцоолол ────────────────────────────────────────────
    def _rng(code_prefix):
        accts = Account.query.filter(
            Account.code.like(code_prefix + '%'), Account.is_active == True
        ).all()
        total = 0
        for a in accts:
            q_dt = db.session.query(_sf.sum(JournalLine.amount)).join(
                Journal, JournalLine.journal_id == Journal.id
            ).filter(
                Journal.status == 'posted',
                Journal.date >= df, Journal.date <= dt,
                JournalLine.debit_account_id == a.id
            ).scalar() or 0
            q_kt = db.session.query(_sf.sum(JournalLine.amount)).join(
                Journal, JournalLine.journal_id == Journal.id
            ).filter(
                Journal.status == 'posted',
                Journal.date >= df, Journal.date <= dt,
                JournalLine.credit_account_id == a.id
            ).scalar() or 0
            total += q_dt - q_kt
        return total

    def _code(code):
        a = Account.query.filter_by(code=code, is_active=True).first()
        if not a: return 0
        q_dt = db.session.query(_sf.sum(JournalLine.amount)).join(
            Journal, JournalLine.journal_id == Journal.id
        ).filter(
            Journal.status == 'posted',
            Journal.date >= df, Journal.date <= dt,
            JournalLine.debit_account_id == a.id
        ).scalar() or 0
        q_kt = db.session.query(_sf.sum(JournalLine.amount)).join(
            Journal, JournalLine.journal_id == Journal.id
        ).filter(
            Journal.status == 'posted',
            Journal.date >= df, Journal.date <= dt,
            JournalLine.credit_account_id == a.id
        ).scalar() or 0
        return q_dt - q_kt

    # ── I. ОРЛОГО ──────────────────────────────────────────────────────────
    rev_service   = -_rng('61')                   # 6110, 6111, 6113 — үйлчилгээний орлого
    rev_other     = -_rng('62')                   # 6200 — бусад үйл ажиллагааны орлого
    rev_financial = -(_code('8410'))              # 8410 — хүүгийн орлого (санхүүгийн)
    rev_total     = rev_service + rev_other + rev_financial

    # ── II. ХАСАГДАХ ЗАРДАЛ ────────────────────────────────────────────────
    exp_cogs      = _rng('7')                     # 7xxx — борлуулалтын шууд зардал
    exp_salary    = _code('8110') + _code('8120') # Цалин
    exp_ndsh      = _code('8111')                 # НД шимтгэл АО
    exp_depr      = _code('8510') + _code('8515') # Элэгдэл
    exp_rent      = _code('8210')                 # Түрээс
    exp_util      = _code('8211') + _code('8212') # Нийтийн үйлчилгээ + холбоо
    exp_travel    = _code('8213') + _code('8214') # Оффис + томилолт
    exp_fin       = _code('8310') + _code('8311') + _code('8312') + _code('8313')  # Санхүүгийн зардал
    exp_other     = _code('8660')                 # Бусад зардал
    exp_total     = (exp_cogs + exp_salary + exp_ndsh + exp_depr +
                     exp_rent + exp_util + exp_travel + exp_fin + exp_other)

    # ── III. ТАТВАР НОГДОХ ОРЛОГО ──────────────────────────────────────────
    profit_before_tax = rev_total - exp_total
    non_deductible    = 0    # Татвар хасагдахгүй зардал — гараар оруулна
    exempt_income     = 0    # Татвараас чөлөөлөгдсөн — гараар оруулна
    taxable_income    = profit_before_tax + non_deductible - exempt_income

    # ── IV. ТАТВАРЫН ТООЦООЛОЛ ─────────────────────────────────────────────
    THRESHOLD = 3_000_000_000   # 3 тэрбум
    RATE_LOW  = 0.10            # 10%
    RATE_HIGH = 0.25            # 25%
    if taxable_income <= 0:
        tax_base_low  = 0
        tax_base_high = 0
        tax_calculated = 0
    elif taxable_income <= THRESHOLD:
        tax_base_low  = taxable_income
        tax_base_high = 0
        tax_calculated = taxable_income * RATE_LOW
    else:
        tax_base_low  = THRESHOLD
        tax_base_high = taxable_income - THRESHOLD
        tax_calculated = THRESHOLD * RATE_LOW + tax_base_high * RATE_HIGH
    tax_relief     = 0    # Хөнгөлөлт — гараар оруулна
    tax_due        = max(0, tax_calculated - tax_relief)

    # ── Урьдчилгаа: 9100 дансаас ────────────────────────────────────────────
    advance_paid  = _code('9100') or 0   # ААТ-ын зардал журналаас

    net_payable   = tax_due - advance_paid
    net_refund    = max(0, -net_payable)
    net_payable   = max(0, net_payable)

    # ── Харьцуулалт: өмнөх жил ─────────────────────────────────────────────
    prev_year = year - 1

    cit = dict(
        year=year, quarter=q_str, period_label=period_label,
        date_from=df, date_to=dt,
        rev_service=rev_service, rev_other=rev_other, rev_financial=rev_financial,
        rev_total=rev_total,
        exp_cogs=exp_cogs, exp_salary=exp_salary, exp_ndsh=exp_ndsh,
        exp_depr=exp_depr, exp_rent=exp_rent, exp_util=exp_util,
        exp_travel=exp_travel, exp_fin=exp_fin, exp_other=exp_other,
        exp_total=exp_total,
        profit_before_tax=profit_before_tax,
        non_deductible=non_deductible, exempt_income=exempt_income,
        taxable_income=taxable_income,
        tax_base_low=tax_base_low, tax_base_high=tax_base_high,
        tax_calculated=tax_calculated, tax_relief=tax_relief, tax_due=tax_due,
        advance_paid=advance_paid, net_payable=net_payable, net_refund=net_refund,
    )
    return render_template('cit_report.html', cit=cit,
                           year=year, quarter=q_str, cur_year=cur_year)


@app.route('/reports/income-flow')
def income_flow_diagram():
    return render_template('income_flow_diagram.html')

@app.route('/reports/income-statement')
def income_statement():
    from sqlalchemy import func as _sf
    report_date = date.today()

    # ── Огноо шүүлт ─────────────────────────────────────────────────────────
    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to', '')
    try:
        df = datetime.strptime(date_from_str, '%Y-%m-%d').date() if date_from_str else None
    except Exception:
        df = None
    try:
        dt = datetime.strptime(date_to_str, '%Y-%m-%d').date() if date_to_str else None
    except Exception:
        dt = None

    def _bal_range(account_id, d_from=None, d_to=None):
        """Огноогоор шүүсэн дансны үлдэгдэл"""
        def _sum(fk_col):
            q = db.session.query(_sf.sum(JournalLine.amount)).join(
                Journal, JournalLine.journal_id == Journal.id
            ).filter(Journal.status == 'posted', fk_col == account_id)
            if d_from: q = q.filter(Journal.date >= d_from)
            if d_to:   q = q.filter(Journal.date <= d_to)
            return q.scalar() or 0
        return _sum(JournalLine.debit_account_id) - _sum(JournalLine.credit_account_id)

    def _subtree_range(root_code, d_from=None, d_to=None):
        accts = Account.query.filter(
            Account.code.like(root_code + '%'), Account.is_active == True
        ).all()
        return sum(_bal_range(a.id, d_from, d_to) for a in accts)

    revenue      = -_subtree_range('6', df, dt)
    cogs         = _subtree_range('7', df, dt)
    gross_profit = revenue - cogs
    ga_expense   = _subtree_range('8', df, dt)
    other_income = 0
    ebit         = gross_profit - ga_expense + other_income
    tax_expense  = _subtree_range('9', df, dt)
    net_profit   = ebit - tax_expense

    odt = dict(
        revenue=revenue, cogs=cogs, gross_profit=gross_profit,
        ga_expense=ga_expense, other_income=other_income,
        ebit=ebit, tax_expense=tax_expense, net_profit=net_profit
    )
    # 2025 comparative (hardcoded)
    odt25 = dict(
        revenue=7380193993.09, cogs=6311339722.93, gross_profit=1068854270.16,
        ga_expense=546081244.58, other_income=1494887.88,
        ebit=521767913.46, tax_expense=52426791.35, net_profit=469341122.11
    )
    # Detail by account group
    rev_detail = []
    for a in Account.query.filter(Account.code.like('6%'), Account.is_active==True).order_by(Account.code).all():
        b = -_bal_range(a.id, df, dt)
        if abs(b) > 0: rev_detail.append({'account': a, 'balance': b})

    exp_detail = []
    for a in Account.query.filter(
        db.or_(Account.code.like('7%'), Account.code.like('8%'), Account.code.like('9%')),
        Account.is_active==True
    ).order_by(Account.code).all():
        b = _bal_range(a.id, df, dt)
        if abs(b) > 0: exp_detail.append({'account': a, 'balance': b})

    return render_template('income_statement.html',
        odt=odt, odt25=odt25,
        rev_detail=rev_detail, exp_detail=exp_detail,
        report_date=report_date,
        date_from=date_from_str, date_to=date_to_str, now=date.today())

@app.route('/reports/cashflow')
def cashflow_report():
    report_date = date.today()
    mode = request.args.get('mode', 'internal')   # 'internal' | 'official'
    from sqlalchemy import func

    ALL_MONTHS = list(range(1, 13))

    # ── Internal cashflow: group by income_cat / expense_cat × month ──
    inc_rows = db.session.query(
        CashTransaction.income_cat,
        CashTransaction.month,
        func.sum(CashTransaction.income).label('amt')
    ).filter(CashTransaction.income > 0, CashTransaction.income_cat != None,
             CashTransaction.income_cat != '')\
     .group_by(CashTransaction.income_cat, CashTransaction.month).all()

    exp_rows = db.session.query(
        CashTransaction.expense_cat,
        CashTransaction.month,
        func.sum(CashTransaction.expense).label('amt')
    ).filter(CashTransaction.expense > 0, CashTransaction.expense_cat != None,
             CashTransaction.expense_cat != '')\
     .group_by(CashTransaction.expense_cat, CashTransaction.month).all()

    # Build {cat: {month: amount}} lookup
    inc_data = {}
    for cat, mon, amt in inc_rows:
        inc_data.setdefault(cat, {})[mon or 0] = float(amt or 0)

    exp_data = {}
    for cat, mon, amt in exp_rows:
        exp_data.setdefault(cat, {})[mon or 0] = float(amt or 0)

    def cat_monthly(data, cat):
        """Return [m1..m12] list for given category"""
        d = data.get(cat, {})
        return [d.get(m, 0) for m in ALL_MONTHS]

    def cats_monthly(data, cats):
        """Sum multiple categories per month"""
        result = [0]*12
        for cat in cats:
            for i, v in enumerate(cat_monthly(data, cat)):
                result[i] += v
        return result

    def row_total(vals): return sum(vals)

    # ── Internal report structure ──
    # INCOME
    r111 = cat_monthly(inc_data, '1.1.1')
    r112 = cat_monthly(inc_data, '1.1.2')
    r113 = cat_monthly(inc_data, '1.1.3')
    r114 = cat_monthly(inc_data, '1.1.4')
    total_inc = [r111[i]+r112[i]+r113[i]+r114[i] for i in range(12)]

    # EXPENSE – contractors
    r121 = cat_monthly(exp_data, '1.2.1')
    r122 = cat_monthly(exp_data, '1.2.2')
    r123 = cat_monthly(exp_data, '1.2.3')
    sub12 = [r121[i]+r122[i]+r123[i] for i in range(12)]

    # EXPENSE – operations
    r211  = cat_monthly(exp_data, '2.1.1')
    r212  = cat_monthly(exp_data, '2.1.2')
    r213  = cat_monthly(exp_data, '2.1.3')
    r214  = cat_monthly(exp_data, '2.1.4')
    r215  = cat_monthly(exp_data, '2.1.5')
    r216  = cat_monthly(exp_data, '2.1.6')
    r217  = cat_monthly(exp_data, '2.1.7')
    r218  = cat_monthly(exp_data, '2.1.8')
    r219  = cat_monthly(exp_data, '2.1.9')
    r2110 = cat_monthly(exp_data, '2.1.10')
    r2111 = cat_monthly(exp_data, '2.1.11')
    r2112 = cat_monthly(exp_data, '2.1.12')
    r2113 = cat_monthly(exp_data, '2.1.13')
    r2114 = cat_monthly(exp_data, '2.1.14')
    r2115 = cat_monthly(exp_data, '2.1.15')
    sub21 = [r211[i]+r212[i]+r213[i]+r214[i]+r215[i]+r216[i]+r217[i]+
             r218[i]+r219[i]+r2110[i]+r2111[i]+r2112[i]+r2113[i]+r2114[i]+r2115[i]
             for i in range(12)]

    # EXPENSE – tax
    r221 = cat_monthly(exp_data, '2.2.1')
    r222 = cat_monthly(exp_data, '2.2.2')
    r223 = cat_monthly(exp_data, '2.2.3')
    r224 = cat_monthly(exp_data, '2.2.4')
    sub22 = [r221[i]+r222[i]+r223[i]+r224[i] for i in range(12)]

    total_exp = [sub12[i]+sub21[i]+sub22[i] for i in range(12)]
    op_net    = [total_inc[i]-total_exp[i] for i in range(12)]

    # INVESTING
    r321 = cat_monthly(exp_data, '3.2.1')
    r322 = cat_monthly(exp_data, '3.2.2')
    r323 = cat_monthly(exp_data, '3.2.3')
    r324 = cat_monthly(exp_data, '3.2.4')
    inv_inc = [0]*12
    inv_exp = [r321[i]+r322[i]+r323[i]+r324[i] for i in range(12)]
    inv_net = [inv_inc[i]-inv_exp[i] for i in range(12)]

    # FINANCING (empty for now)
    fin_inc = [0]*12
    fin_exp = [0]*12
    fin_net = [0]*12

    # RELATED PARTIES (section 5)
    r512 = cat_monthly(inc_data, '5.1.2') if '5.1.2' in inc_data else cat_monthly(exp_data, '5.1.2') if '5.1.2' in exp_data else [0]*12
    r513 = cat_monthly(inc_data, '5.1.3') if '5.1.3' in inc_data else [0]*12
    sub51 = [r512[i]+r513[i] for i in range(12)]
    r521 = cat_monthly(exp_data, '5.2.1') if '5.2.1' in exp_data else [0]*12
    r522 = cat_monthly(exp_data, '5.2.2') if '5.2.2' in exp_data else [0]*12
    r523 = cat_monthly(exp_data, '5.2.3') if '5.2.3' in exp_data else [0]*12
    sub52 = [r521[i]+r522[i]+r523[i] for i in range(12)]
    sec5_net = [sub51[i]-sub52[i] for i in range(12)]

    # Net total
    OPEN_CASH = 173541873.71
    net_total = [op_net[i]+inv_net[i]+fin_net[i]+sec5_net[i] for i in range(12)]
    closing   = []
    bal = OPEN_CASH
    for v in net_total:
        bal += v
        closing.append(bal)
    opening_by_month = [OPEN_CASH] + closing[:-1]

    internal = dict(
        months=ALL_MONTHS,
        # income
        r111=r111, r112=r112, r113=r113, r114=r114, total_inc=total_inc,
        # contractor exp
        r121=r121, r122=r122, r123=r123, sub12=sub12,
        # ops exp
        r211=r211, r212=r212, r213=r213, r214=r214, r215=r215,
        r216=r216, r217=r217, r218=r218, r219=r219,
        r2110=r2110, r2111=r2111, r2112=r2112, r2113=r2113, r2114=r2114, r2115=r2115,
        sub21=sub21,
        # tax
        r221=r221, r222=r222, r223=r223, r224=r224, sub22=sub22,
        # totals
        total_exp=total_exp, op_net=op_net,
        # investing
        r321=r321, r322=r322, r323=r323, r324=r324,
        inv_inc=inv_inc, inv_exp=inv_exp, inv_net=inv_net,
        # financing
        fin_inc=fin_inc, fin_exp=fin_exp, fin_net=fin_net,
        # related
        r512=r512, r513=r513, sub51=sub51,
        r521=r521, r522=r522, r523=r523, sub52=sub52, sec5_net=sec5_net,
        # net
        net_total=net_total,
        open_cash=OPEN_CASH,
        opening_by_month=opening_by_month,
        closing=closing,
    )

    # ── Official format (existing) ──
    monthly_q = db.session.query(
        CashTransaction.month,
        func.sum(CashTransaction.income).label('income'),
        func.sum(CashTransaction.expense).label('expense')
    ).group_by(CashTransaction.month).order_by(CashTransaction.month).all()

    op_income_total  = db.session.query(func.sum(CashTransaction.income)).scalar() or 0
    op_expense_total = db.session.query(func.sum(CashTransaction.expense)).scalar() or 0

    mgt = dict(op_income=op_income_total, op_expense=op_expense_total,
               op_net=op_income_total-op_expense_total, monthly=monthly_q)
    mgt25 = dict(
        op_income=8142842883.60, op_expense=7835279640.86, op_net=307563242.74,
        inv_income=40418635.88, inv_expense=199927700, inv_net=-159509064.12,
        fin_income=144000000, fin_expense=144200000, fin_net=-200000,
        total_net=147854178.62,
        open_cash=25878883.58, close_cash=173733062.20
    )

    return render_template('cashflow_report.html',
                           mode=mode, internal=internal,
                           mgt=mgt, mgt25=mgt25,
                           report_date=report_date)

@app.route('/reports/equity')
def equity_report():
    report_date = date.today()
    # 2025 тайлан + 2026 өөрчлөлт
    net_profit_2026 = -(subtree_bal('5') - (-522004441.13))
    eq = dict(
        open_capital=20000, open_retained=521984441.13,
        open_total=522004441.13,
        net_profit=net_profit_2026,
        close_capital=20000,
        close_retained=521984441.13 + net_profit_2026,
        close_total=522004441.13 + net_profit_2026
    )
    return render_template('equity_report.html', eq=eq, report_date=report_date)

@app.route('/api/account-drilldown')
def api_account_drilldown():
    """Данс + хугацааны journal lines задаргаа — JSON"""
    acc_id    = request.args.get('acc_id',    type=int)
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to',   '')
    side      = request.args.get('side', 'both')   # dt | kt | both

    if not acc_id:
        return jsonify({'error': 'acc_id required'}), 400

    acc = Account.query.get_or_404(acc_id)

    filt = [Journal.status == 'posted']
    if date_from:
        try:
            filt.append(Journal.date >= datetime.strptime(date_from, '%Y-%m-%d').date())
        except: pass
    if date_to:
        try:
            filt.append(Journal.date <= datetime.strptime(date_to, '%Y-%m-%d').date())
        except: pass

    lines = []
    # ДТ мөрүүд
    if side in ('dt', 'both'):
        dt_lines = (db.session.query(JournalLine, Journal)
            .join(Journal, JournalLine.journal_id == Journal.id)
            .filter(JournalLine.debit_account_id == acc_id, *filt)
            .order_by(Journal.date, Journal.number)
            .all())
        for jl, j in dt_lines:
            kt_acc = Account.query.get(jl.credit_account_id)
            lines.append({
                'id': jl.id, 'side': 'dt',
                'date': j.date.strftime('%Y-%m-%d') if j.date else '',
                'journal_no': j.number or '',
                'description': jl.description or j.description or '',
                'amount': jl.amount,
                'counterpart_code': kt_acc.code if kt_acc else '',
                'counterpart_name': kt_acc.name if kt_acc else '',
            })

    # КТ мөрүүд
    if side in ('kt', 'both'):
        kt_lines = (db.session.query(JournalLine, Journal)
            .join(Journal, JournalLine.journal_id == Journal.id)
            .filter(JournalLine.credit_account_id == acc_id, *filt)
            .order_by(Journal.date, Journal.number)
            .all())
        for jl, j in kt_lines:
            dt_acc = Account.query.get(jl.debit_account_id)
            lines.append({
                'id': jl.id, 'side': 'kt',
                'date': j.date.strftime('%Y-%m-%d') if j.date else '',
                'journal_no': j.number or '',
                'description': jl.description or j.description or '',
                'amount': jl.amount,
                'counterpart_code': dt_acc.code if dt_acc else '',
                'counterpart_name': dt_acc.name if dt_acc else '',
            })

    lines.sort(key=lambda x: x['date'])
    total = sum(l['amount'] for l in lines)

    return jsonify({
        'acc_code': acc.code,
        'acc_name': acc.name,
        'side': side,
        'date_from': date_from,
        'date_to': date_to,
        'count': len(lines),
        'total': total,
        'lines': lines,
    })


@app.route('/reports/trial-balance')
def trial_balance():
    from sqlalchemy import func as sqlfunc

    # Огноо шүүлт — default: тухайн оны 1 дүгээр сараас өнөөдөр хүртэл
    import datetime as _dt
    _today = _dt.date.today()
    _default_from = f'{_today.year}-01-01'
    _default_to   = _today.strftime('%Y-%m-%d')
    date_from_str = request.args.get('date_from', _default_from)
    date_to_str   = request.args.get('date_to',   _default_to)
    show_zero     = request.args.get('show_zero', '0') == '1'

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
    except Exception:
        date_from = None
    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except Exception:
        date_to = None

    # Бүх идэвхтэй данс
    accounts = Account.query.filter_by(is_active=True).order_by(Account.code).all()

    def get_period_amounts(account_id, d_from, d_to):
        """Тодорхой хугацааны ДТ/КТ нийлбэр"""
        filt = [Journal.status == 'posted']
        if d_from: filt.append(Journal.date >= d_from)
        if d_to:   filt.append(Journal.date <= d_to)
        j_ids = db.session.query(Journal.id).filter(*filt).subquery()

        dt = db.session.query(sqlfunc.sum(JournalLine.amount)).filter(
            JournalLine.debit_account_id  == account_id,
            JournalLine.journal_id.in_(j_ids)
        ).scalar() or 0
        kt = db.session.query(sqlfunc.sum(JournalLine.amount)).filter(
            JournalLine.credit_account_id == account_id,
            JournalLine.journal_id.in_(j_ids)
        ).scalar() or 0
        return dt, kt

    rows = []
    tot_ob_dt = tot_ob_kt = 0
    tot_dt = tot_kt = 0
    tot_cl_dt = tot_cl_kt = 0

    for acc in accounts:
        # Эхний үлдэгдэл — date_from өмнөх бүх гүйлгээ
        if date_from:
            ob_dt, ob_kt = get_period_amounts(acc.id, None,
                                              date(date_from.year, date_from.month, date_from.day - 1)
                                              if date_from.day > 1 else date_from)
            # Энгийн арга: date_from-аас 1 өдөр өмнөх хүртэлх
            from datetime import timedelta
            prev = date_from - timedelta(days=1)
            ob_dt, ob_kt = get_period_amounts(acc.id, None, prev)
        else:
            ob_dt, ob_kt = 0, 0

        # Тухайн хугацааны гүйлгээ
        p_dt, p_kt = get_period_amounts(acc.id, date_from, date_to)

        # Эхний үлдэгдэл нет
        ob_net = ob_dt - ob_kt
        ob_show_dt = ob_net if ob_net > 0 else 0
        ob_show_kt = -ob_net if ob_net < 0 else 0

        # Эцсийн үлдэгдэл
        cl_net = ob_net + p_dt - p_kt
        cl_show_dt = cl_net if cl_net > 0 else 0
        cl_show_kt = -cl_net if cl_net < 0 else 0

        # Тэг данс алгасах
        if not show_zero and ob_net == 0 and p_dt == 0 and p_kt == 0:
            continue

        rows.append(dict(
            acc=acc,
            ob_dt=ob_show_dt, ob_kt=ob_show_kt,
            p_dt=p_dt,        p_kt=p_kt,
            cl_dt=cl_show_dt, cl_kt=cl_show_kt,
        ))
        tot_ob_dt += ob_show_dt; tot_ob_kt += ob_show_kt
        tot_dt    += p_dt;       tot_kt    += p_kt
        tot_cl_dt += cl_show_dt; tot_cl_kt += cl_show_kt

    # ── Бүлгийн нийлбэр тооцоолох ──────────────────────────────────────────
    GROUP_LABELS = {
        '11':'Мөнгө, мөнгөн хөрөнгө', '12':'Авлага', '13':'Урьдчилж төлсөн',
        '14':'Бараа материал',          '15':'Үндсэн хөрөнгө (нэгтгэл)', '16':'Биет бус хөрөнгө',
        '20':'Үндсэн хөрөнгө (задаргаа)',
        '21':'Урт хугацаат авлага',
        '31':'Богино хугацаат өглөг',  '32':'Бусад богино хугацаат өр',
        '33':'Татварын өр',             '34':'НДШ өр',
        '40':'Урт хугацаат зээл',       '41':'Бусад урт хугацаат өр',
        '51':'Дүрмийн сан',            '52':'Нэмж төлөгдсөн капитал',
        '53':'Хуримтлагдсан ашиг',     '61':'Борлуулалтын орлого',
        '62':'Бусад орлого',            '72':'Борлуулалтын өртөг',
        '73':'Борлуулалтын зардал',     '74':'Ерөнхий удирдлагын зардал',
        '75':'Санхүүгийн зардал',      '91':'Татварын зардал',
    }
    SECTION_LABELS = {
        '1':'1. ХӨРӨНГӨ', '2':'2. УРТ ХУГАЦААТ ХӨРӨНГӨ',
        '3':'3. БОГИНО ХУГАЦААТ ӨР ТӨЛБӨР', '4':'4. УРТ ХУГАЦААТ ӨР ТӨЛБӨР',
        '5':'5. ЭЗДИЙН ӨМЧ', '6':'6. ОРЛОГО',
        '7':'7. ЗАРДАЛ', '8':'8. ЗАРДАЛ', '9':'9. ТАТВАР',
    }

    def _sum_buf(buf):
        return dict(
            ob_dt=sum(r['ob_dt'] for r in buf), ob_kt=sum(r['ob_kt'] for r in buf),
            p_dt=sum(r['p_dt'] for r in buf),   p_kt=sum(r['p_kt'] for r in buf),
            cl_dt=sum(r['cl_dt'] for r in buf), cl_kt=sum(r['cl_kt'] for r in buf),
        )

    data_rows_count = len(rows)
    final_rows  = []
    cur_section = [None]
    cur_prefix  = [None]
    grp_buf     = []
    sec_buf     = []

    def _flush_grp():
        if not grp_buf:
            return
        pf  = grp_buf[0]['acc'].code[:2]
        lbl = GROUP_LABELS.get(pf, f'{pf}xx')
        final_rows.append({'type': 'header', 'label': f'{pf}xx — {lbl}'})
        for r in grp_buf[:]:
            final_rows.append(r)
        s = _sum_buf(grp_buf); s['type'] = 'subtotal'; s['label'] = f'{pf}xx нийт'
        final_rows.append(s)
        sec_buf.extend(grp_buf[:])
        grp_buf.clear()

    def _flush_sec():
        if not sec_buf:
            return
        lbl = SECTION_LABELS.get(cur_section[0] or '', f'{cur_section[0]}xxx')
        s   = _sum_buf(sec_buf); s['type'] = 'section'; s['label'] = lbl + ' — НИЙТ'
        final_rows.append(s)
        final_rows.append({'type': 'spacer'})
        sec_buf.clear()

    for row in rows:
        code = row['acc'].code
        pf   = code[:2] if len(code) >= 2 else code
        sec  = code[:1]
        if sec != cur_section[0]:
            _flush_grp(); _flush_sec()
            cur_section[0] = sec; cur_prefix[0] = None
        if pf != cur_prefix[0]:
            _flush_grp(); cur_prefix[0] = pf
        grp_buf.append(row)
    _flush_grp(); _flush_sec()

    return render_template('trial_balance.html',
                           rows=final_rows,
                           data_rows_count=data_rows_count,
                           date_from=date_from_str, date_to=date_to_str,
                           show_zero=show_zero,
                           tot_ob_dt=tot_ob_dt, tot_ob_kt=tot_ob_kt,
                           tot_dt=tot_dt,       tot_kt=tot_kt,
                           tot_cl_dt=tot_cl_dt, tot_cl_kt=tot_cl_kt)

@app.route('/reports/ledger')
def ledger():
    account_id = request.args.get('account_id', type=int)
    date_from  = request.args.get('date_from', '')
    date_to    = request.args.get('date_to', '')
    accounts   = Account.query.filter_by(is_active=True).order_by(Account.code).all()
    selected   = None
    lines      = []
    opening_balance = 0
    partner_by_journal = {}

    if account_id:
        selected = Account.query.get(account_id)
        q = JournalLine.query.filter(
            db.or_(
                JournalLine.debit_account_id == account_id,
                JournalLine.credit_account_id == account_id
            )
        ).join(Journal).filter(Journal.status == 'posted')

        # Opening balance = everything before date_from
        if date_from:
            df = datetime.strptime(date_from, '%Y-%m-%d').date()
            before = q.filter(Journal.date < df).all()
            for bl in before:
                if bl.debit_account_id == account_id:
                    opening_balance += bl.amount
                else:
                    opening_balance -= bl.amount
            q = q.filter(Journal.date >= df)

        if date_to:
            dt = datetime.strptime(date_to, '%Y-%m-%d').date()
            q = q.filter(Journal.date <= dt)

        lines = q.order_by(Journal.date, Journal.id).all()

        # journal_id → partner_name lookup (CashTransaction-аас)
        jids = list({l.journal_id for l in lines if l.journal_id})
        partner_by_journal = {}
        if jids:
            cash_rows = CashTransaction.query.filter(
                CashTransaction.journal_id.in_(jids)
            ).with_entities(
                CashTransaction.journal_id,
                CashTransaction.partner_name,
                CashTransaction.master_name,
            ).all()
            # Нэг journal-д олон гүйлгээ байвал өвөрмөц нэрүүд нэгтгэнэ (max 2)
            from collections import defaultdict
            _tmp = defaultdict(set)
            for cr in cash_rows:
                nm = (cr.partner_name or cr.master_name or '').strip()
                if nm:
                    _tmp[cr.journal_id].add(nm)
            for jid, names in _tmp.items():
                nl = sorted(names)
                if len(nl) > 2:
                    partner_by_journal[jid] = f"{nl[0]}, {nl[1]} (+{len(nl)-2})"
                else:
                    partner_by_journal[jid] = ', '.join(nl)

    return render_template('ledger.html', accounts=accounts, selected=selected,
                           lines=lines, opening_balance=opening_balance,
                           date_from=date_from, date_to=date_to,
                           partner_by_journal=partner_by_journal)

# ===================== БАРИМТ ТУЛГАЛТ =====================

@app.route('/recon/barimt-bank')
def barimt_bank_recon():
    """eBarimt ↔ Банкны хуулга тулгалт — журнал бичилтгүй"""
    from sqlalchemy import func as sqf

    sel_month  = request.args.get('month', '', type=str)
    sel_type   = request.args.get('vtype', 'out')   # out=борлуулалт, in=худалдан авалт
    sel_partner= request.args.get('partner_id', '', type=str)

    # ── eBarimt нэгтгэл харилцагчаар ──
    vat_q = db.session.query(
        VatRecord.partner_id,
        VatRecord.partner_name,
        sqf.sum(VatRecord.total_amount).label('vat_total'),
        sqf.sum(VatRecord.vat_amount).label('vat_noat'),
        sqf.count(VatRecord.id).label('vat_cnt'),
    ).filter(VatRecord.type == sel_type)

    if sel_month:
        vat_q = vat_q.filter(VatRecord.month == int(sel_month))
    if sel_partner:
        vat_q = vat_q.filter(VatRecord.partner_id == int(sel_partner))

    vat_rows = vat_q.group_by(VatRecord.partner_id, VatRecord.partner_name).all()

    # partner_id → vat дүн
    vat_by_pid  = {}
    vat_noname  = []   # partner_id байхгүй
    for row in vat_rows:
        if row.partner_id:
            vat_by_pid[row.partner_id] = row
        else:
            vat_noname.append(row)

    # ── Банкны хуулга нэгтгэл харилцагчаар ──
    if sel_type == 'out':
        # борлуулалт → орлого
        cash_col = CashTransaction.income
    else:
        # худалдан авалт → зарлага
        cash_col = CashTransaction.expense

    cash_q = db.session.query(
        CashTransaction.partner_id,
        sqf.sum(cash_col).label('bank_total'),
        sqf.count(CashTransaction.id).label('bank_cnt'),
    ).filter(cash_col > 0)

    if sel_month:
        cash_q = cash_q.filter(CashTransaction.month == int(sel_month))
    if sel_partner:
        cash_q = cash_q.filter(CashTransaction.partner_id == int(sel_partner))

    cash_rows = cash_q.group_by(CashTransaction.partner_id).all()
    bank_by_pid = {r.partner_id: r for r in cash_rows if r.partner_id}

    # ── Нэгтгэж хүснэгт үүсгэнэ ──
    all_pids = set(list(vat_by_pid.keys()) + list(bank_by_pid.keys()))
    partners_map = {p.id: p for p in Partner.query.filter(Partner.id.in_(all_pids)).all()}

    rows = []
    for pid in sorted(all_pids):
        p    = partners_map.get(pid)
        vr   = vat_by_pid.get(pid)
        br   = bank_by_pid.get(pid)
        vat_amt  = vr.vat_total  if vr else 0
        bank_amt = br.bank_total if br else 0
        diff     = bank_amt - vat_amt
        rows.append({
            'partner_id':   pid,
            'partner_name': p.name if p else '—',
            'vat_total':    vat_amt  or 0,
            'vat_cnt':      vr.vat_cnt   if vr else 0,
            'bank_total':   bank_amt or 0,
            'bank_cnt':     br.bank_cnt  if br else 0,
            'diff':         diff,
            'status':       'match' if abs(diff) < 1 else ('bank_more' if diff > 0 else 'vat_more'),
        })

    # Эрэмбэлэх: зөрүүтэй эхлээд
    rows.sort(key=lambda r: abs(r['diff']), reverse=True)

    # Нийт
    total_vat  = sum(r['vat_total']  for r in rows)
    total_bank = sum(r['bank_total'] for r in rows)
    total_diff = total_bank - total_vat

    # Сарын нэгтгэл (шүүлтгүй)
    monthly_vat = db.session.query(
        VatRecord.month,
        sqf.sum(VatRecord.total_amount),
        sqf.count(VatRecord.id),
    ).filter(VatRecord.type == sel_type).group_by(VatRecord.month).order_by(VatRecord.month).all()

    monthly_bank = db.session.query(
        CashTransaction.month,
        sqf.sum(cash_col),
        sqf.count(CashTransaction.id),
    ).filter(cash_col > 0).group_by(CashTransaction.month).order_by(CashTransaction.month).all()

    mv = {r[0]: (r[1] or 0, r[2]) for r in monthly_vat}
    mb = {r[0]: (r[1] or 0, r[2]) for r in monthly_bank}
    all_months = sorted(set(list(mv.keys()) + list(mb.keys())))
    monthly = []
    for m in all_months:
        va, vc = mv.get(m, (0, 0))
        ba, bc = mb.get(m, (0, 0))
        monthly.append({'m': m, 'vat': va, 'bank': ba, 'diff': ba - va,
                        'vc': vc, 'bc': bc})

    # Partner жагсаалт (шүүлтэнд)
    partners_list = Partner.query.filter(Partner.id.in_(all_pids)).order_by(Partner.name).all()

    return render_template('barimt_bank_recon.html',
        rows=rows, total_vat=total_vat, total_bank=total_bank, total_diff=total_diff,
        monthly=monthly, sel_month=sel_month, sel_type=sel_type,
        sel_partner=sel_partner, partners_list=partners_list,
        vat_noname=vat_noname,
    )


# --- VAT ---
@app.route('/vat')
def vat():
    from sqlalchemy import func
    vtype  = request.args.get('type', '')
    month  = request.args.get('month', type=int)
    search = request.args.get('q', '').strip()

    q = VatRecord.query
    if vtype:  q = q.filter_by(type=vtype)
    if month:  q = q.filter_by(month=month)
    if search: q = q.filter(db.or_(
        VatRecord.partner_name.ilike(f'%{search}%'),
        VatRecord.partner_register.ilike(f'%{search}%'),
        VatRecord.ddtd.ilike(f'%{search}%')
    ))
    records = q.order_by(VatRecord.date.desc()).all()

    # Totals
    out_vat = db.session.query(func.sum(VatRecord.vat_amount)).filter_by(type='out').scalar() or 0
    in_vat  = db.session.query(func.sum(VatRecord.vat_amount)).filter_by(type='in').scalar() or 0
    out_total = db.session.query(func.sum(VatRecord.total_amount)).filter_by(type='out').scalar() or 0
    in_total  = db.session.query(func.sum(VatRecord.total_amount)).filter_by(type='in').scalar() or 0
    net_vat   = out_vat - in_vat

    # Monthly breakdown
    monthly = db.session.query(
        VatRecord.month,
        VatRecord.type,
        func.sum(VatRecord.vat_amount).label('vat'),
        func.sum(VatRecord.total_amount).label('total'),
        func.count().label('cnt')
    ).group_by(VatRecord.month, VatRecord.type)\
     .order_by(VatRecord.month).all()

    months_list = db.session.query(VatRecord.month).filter(VatRecord.month != None)\
                    .distinct().order_by(VatRecord.month).all()
    months_list = [m[0] for m in months_list]

    return render_template('vat.html', records=records,
                           out_vat=out_vat, in_vat=in_vat, net_vat=net_vat,
                           out_total=out_total, in_total=in_total,
                           monthly=monthly, months_list=months_list,
                           sel_type=vtype, sel_month=month, search=search)

@app.route('/vat/<int:vat_id>/toggle-type', methods=['POST'])
def vat_toggle_type(vat_id):
    """eBarimt-ын төрлийг солих: out ↔ in"""
    v = VatRecord.query.get_or_404(vat_id)
    v.type = 'in' if v.type == 'out' else 'out'
    db.session.commit()
    return jsonify({'ok': True, 'new_type': v.type})

@app.route('/vat/unmatched', methods=['GET', 'POST'])
def vat_unmatched():
    """Тааралтгүй харилцагчийн нэрүүд — Partner тулгах / нэмэх."""
    from name_match import match_partner
    from collections import defaultdict
    from sqlalchemy import func as sqf

    if request.method == 'POST':
        action = request.form.get('action', '')

        # ── 1. Одоо байгаа Partner-т тулгах ──────────────────────────────
        if action == 'assign':
            vat_name   = request.form.get('vat_name', '').strip()
            partner_id = request.form.get('partner_id', type=int)
            vat_reg    = request.form.get('vat_register', '').strip()
            save_reg   = request.form.get('save_reg') == '1'

            if vat_name and partner_id:
                # Partner-ийн регистрийг шинэчлэх
                p = db.session.get(Partner, partner_id)
                if p and save_reg and vat_reg and not p.register:
                    p.register = vat_reg

                # Энэ нэртэй бүх VatRecord-ийг тулгах
                recs = VatRecord.query.filter(
                    VatRecord.partner_id == None,
                    VatRecord.partner_name == vat_name
                ).all()
                for r in recs:
                    r.partner_id = partner_id
                db.session.commit()
                flash(f'"{vat_name}" → [{p.name if p else partner_id}] — {len(recs)} баримт тулгагдлаа.', 'success')

        # ── 2. Шинэ Partner нэмэх ──────────────────────────────────────
        elif action == 'add_new':
            vat_name = request.form.get('vat_name', '').strip()
            vat_reg  = request.form.get('vat_register', '').strip()
            if vat_name:
                existing = Partner.query.filter(
                    db.func.upper(Partner.name) == vat_name.upper()
                ).first()
                if not existing:
                    p = Partner(
                        name=vat_name,
                        register=vat_reg or None,
                        type='both',
                        is_active=True
                    )
                    db.session.add(p)
                    db.session.flush()
                    # Тулгах
                    recs = VatRecord.query.filter(
                        VatRecord.partner_id == None,
                        VatRecord.partner_name == vat_name
                    ).all()
                    for r in recs:
                        r.partner_id = p.id
                    db.session.commit()
                    flash(f'Шинэ харилцагч [{vat_name}] нэмэгдлээ, {len(recs)} баримт тулгагдлаа.', 'success')
                else:
                    flash(f'[{vat_name}] аль хэдийн байна (id={existing.id}).', 'warning')

        # ── 3. Бөөнөөр шинэ Partner нэмэх ────────────────────────────
        elif action == 'bulk_add':
            names = request.form.getlist('add_names')
            added = matched = 0
            for name in names:
                name = name.strip()
                if not name:
                    continue
                if not Partner.query.filter(db.func.upper(Partner.name) == name.upper()).first():
                    p = Partner(name=name, type='both', is_active=True)
                    db.session.add(p)
                    added += 1
            db.session.commit()
            if added:
                partners_all = Partner.query.filter_by(is_active=True).all()
                recs = VatRecord.query.filter(VatRecord.partner_id == None).all()
                for r in recs:
                    pm = match_partner(r.partner_name or '', r.partner_register or '', partners_all)
                    if pm:
                        r.partner_id = pm.id
                        matched += 1
                db.session.commit()
            flash(f'{added} харилцагч нэмэгдлээ, {matched} баримт тулгагдлаа.', 'success')

        return redirect(url_for('vat_unmatched'))

    # ── GET: тааралтгүй нэрүүдийг бүлэглэнэ ──────────────────────────────
    recs = VatRecord.query.filter(VatRecord.partner_id == None).all()

    # Нэр бүрийн: count, total_amount, register жагсаалт, type
    groups = defaultdict(lambda: {'cnt': 0, 'total': 0.0, 'regs': set(), 'types': set()})
    for r in recs:
        name = (r.partner_name or '').strip()
        if not name:
            continue
        groups[name]['cnt']   += 1
        groups[name]['total'] += (r.total_amount or 0)
        if r.partner_register:
            groups[name]['regs'].add(r.partner_register.strip())
        groups[name]['types'].add(r.type or '')

    unmatched = []
    for name, info in sorted(groups.items(), key=lambda x: -x[1]['cnt']):
        regs = list(info['regs'])
        unmatched.append({
            'name':  name,
            'cnt':   info['cnt'],
            'total': info['total'],
            'reg':   regs[0] if regs else '',
            'type':  ','.join(sorted(info['types'])),
        })

    # Partner жагсаалт (тулгахад ашиглана)
    partners_list = Partner.query.filter_by(is_active=True).order_by(Partner.name).all()
    total_amount  = sum(g['total'] for g in unmatched)

    return render_template('vat_unmatched.html',
        unmatched=unmatched,
        total_recs=len(recs),
        total_amount=total_amount,
        partners_list=partners_list,
    )


@app.route('/vat/reconcile', methods=['POST'])
def vat_reconcile():
    """Бүх тааралтгүй VatRecord-уудыг дахин тулгана."""
    from name_match import match_partner
    partners = Partner.query.filter_by(is_active=True).all()
    recs = VatRecord.query.filter(VatRecord.partner_id == None).all()
    matched = 0
    for r in recs:
        p = match_partner(r.partner_name or '', r.partner_register or '', partners)
        if p:
            r.partner_id = p.id
            matched += 1
    db.session.commit()
    flash(f'Дахин тулгалт: {matched} / {len(recs)} VatRecord тулгалт олдлоо.', 'success')
    return redirect(url_for('vat_unmatched'))


# ── eBarimt-аас нэхэмжлэл үүсгэх (нэг тус бүр) ─────────────────────────────
@app.route('/vat/<int:vat_id>/make-invoice', methods=['POST'])
def vat_make_invoice(vat_id):
    """VatRecord (type='out')-аас Receivable + RECV-INV журнал үүсгэнэ."""
    v = VatRecord.query.get_or_404(vat_id)
    if v.type != 'out':
        return jsonify({'ok': False, 'error': 'Зөвхөн борлуулалтын eBarimt-аас нэхэмжлэл үүсгэнэ'}), 400
    if v.parent_ddtd:
        return jsonify({'ok': False, 'error': f'Энэ баримт нь хаалтын баримт (Толгой ДДТД: {v.parent_ddtd}). Орлого биш — нэхэмжлэл үүсгэхгүй.'}), 400
    if v.receivable_id:
        r = Receivable.query.get(v.receivable_id)
        return jsonify({'ok': False, 'error': f'Аль хэдийн нэхэмжлэл үүссэн: {r.invoice_no if r else v.receivable_id}'}), 400
    if not v.partner_id:
        return jsonify({'ok': False, 'error': 'Харилцагч тулгагдаагүй байна'}), 400

    from sqlalchemy import func as sqlfunc
    # Нэхэмжлэлийн дугаар: INV2NNNN (eBarimt ДДТД-г суурилуулна)
    last = Receivable.query.order_by(Receivable.id.desc()).first()
    next_num = (last.id + 1) if last else 1
    ddtd_short = (v.ddtd or '').replace(' ', '')[-6:] if v.ddtd else ''
    invoice_no = f'INV{ddtd_short or next_num}'
    # Давхардал байвал дугаар нэмнэ
    if Receivable.query.filter_by(invoice_no=invoice_no).first():
        invoice_no = f'INV{next_num}'

    r = Receivable(
        date=v.date,
        due_date=None,
        invoice_no=invoice_no,
        partner_id=v.partner_id,
        description=f'eBarimt {v.ddtd or ""} — {v.partner_name or ""}',
        amount=v.total_amount,
        status='open'
    )
    db.session.add(r)
    db.session.flush()

    # RECV-INV журнал: DT 1210 / KT 6110 + (НӨАТ байвал DT 1210 / KT 3152)
    acc_1210 = Account.query.filter_by(code='1210').first()
    acc_6110 = Account.query.filter_by(code='6110').first()
    acc_3152 = Account.query.filter_by(code='3152').first()

    if acc_1210 and acc_6110:
        j = Journal(
            date=v.date,
            number=f'RECV-{invoice_no}',
            description=f'eBarimt {v.ddtd or ""} — {v.partner_name or ""}',
            reference=invoice_no,
            status='posted'
        )
        db.session.add(j)
        db.session.flush()
        # Орлогын мөр (НӨАТ-гүй дүн)
        db.session.add(JournalLine(
            journal_id=j.id,
            debit_account_id=acc_1210.id,
            credit_account_id=acc_6110.id,
            amount=v.amount,
            description=f'Тээврийн орлого — {v.partner_name or ""}'
        ))
        # НӨАТ мөр
        if v.vat_amount and v.vat_amount > 0 and acc_3152:
            db.session.add(JournalLine(
                journal_id=j.id,
                debit_account_id=acc_1210.id,
                credit_account_id=acc_3152.id,
                amount=v.vat_amount,
                description='НӨАТ 10%'
            ))

    # VatRecord-т холбоно
    v.receivable_id = r.id
    db.session.commit()

    return jsonify({
        'ok': True,
        'invoice_no': invoice_no,
        'receivable_id': r.id,
        'amount': v.total_amount,
    })


# ── eBarimt-аас нэхэмжлэл БАГЦААР үүсгэх ────────────────────────────────────
@app.route('/vat/bulk-make-invoices', methods=['POST'])
def vat_bulk_make_invoices():
    """partner_id тулгагдсан, нэхэмжлэлгүй бүх OUT eBarimt-аас нэхэмжлэл үүсгэнэ."""
    from sqlalchemy import func as sqlfunc
    data      = request.get_json() or {}
    month_fil = data.get('month')     # optional
    pid_fil   = data.get('partner_id')

    q = VatRecord.query.filter(
        VatRecord.type == 'out',
        VatRecord.partner_id != None,
        VatRecord.receivable_id == None,
        db.or_(VatRecord.parent_ddtd == None, VatRecord.parent_ddtd == '')  # Зөвхөн анхны баримт
    )
    if month_fil:
        q = q.filter(VatRecord.month == int(month_fil))
    if pid_fil:
        q = q.filter(VatRecord.partner_id == int(pid_fil))
    vats = q.order_by(VatRecord.date).all()

    acc_1210 = Account.query.filter_by(code='1210').first()
    acc_6110 = Account.query.filter_by(code='6110').first()
    acc_3152 = Account.query.filter_by(code='3152').first()

    last = Receivable.query.order_by(Receivable.id.desc()).first()
    next_num = (last.id + 1) if last else 1

    created = 0
    skipped = 0
    for v in vats:
        ddtd_short = (v.ddtd or '').replace(' ', '')[-6:] if v.ddtd else ''
        invoice_no = f'INV{ddtd_short or next_num}'
        if Receivable.query.filter_by(invoice_no=invoice_no).first():
            invoice_no = f'INV{next_num}'

        r = Receivable(
            date=v.date,
            due_date=None,
            invoice_no=invoice_no,
            partner_id=v.partner_id,
            description=f'eBarimt {v.ddtd or ""} — {v.partner_name or ""}',
            amount=v.total_amount,
            status='open'
        )
        db.session.add(r)
        db.session.flush()

        if acc_1210 and acc_6110:
            j = Journal(
                date=v.date,
                number=f'RECV-{invoice_no}',
                description=f'eBarimt {v.ddtd or ""} — {v.partner_name or ""}',
                reference=invoice_no,
                status='posted'
            )
            db.session.add(j)
            db.session.flush()
            db.session.add(JournalLine(
                journal_id=j.id,
                debit_account_id=acc_1210.id,
                credit_account_id=acc_6110.id,
                amount=v.amount,
                description=f'Тээврийн орлого — {v.partner_name or ""}'
            ))
            if v.vat_amount and v.vat_amount > 0 and acc_3152:
                db.session.add(JournalLine(
                    journal_id=j.id,
                    debit_account_id=acc_1210.id,
                    credit_account_id=acc_3152.id,
                    amount=v.vat_amount,
                    description='НӨАТ 10%'
                ))

        v.receivable_id = r.id
        next_num += 1
        created += 1

    db.session.commit()
    return jsonify({'ok': True, 'created': created, 'skipped': skipped, 'total': len(vats)})


@app.route('/vat/add', methods=['GET', 'POST'])
def add_vat():
    if request.method == 'POST':
        amount = float(request.form.get('amount', 0))
        vat_pct = float(request.form.get('vat_pct', 10)) / 100
        vat_amount = round(amount * vat_pct, 2)
        pid_raw = request.form.get('partner_id', '').strip()
        pid = int(pid_raw) if pid_raw else None
        # Харилцагч сонгосон бол нэр/регистр автоматаар авах
        pname = request.form.get('partner_name', '').strip()
        preg  = request.form.get('partner_register', '').strip()
        if pid:
            pp = Partner.query.get(pid)
            if pp:
                pname = pp.name
                preg  = pp.register or preg
        rec = VatRecord(
            date=datetime.strptime(request.form.get('date'), '%Y-%m-%d').date(),
            type=request.form.get('type'),
            invoice_no=request.form.get('invoice_no', ''),
            partner_id=pid,
            partner_name=pname,
            partner_register=preg,
            amount=amount,
            vat_amount=vat_amount,
            total_amount=amount + vat_amount,
            ebarimt_sent=request.form.get('ebarimt_sent') == 'on',
            source='Гар',
        )
        db.session.add(rec)
        db.session.commit()
        flash('НӨАТ бүртгэл амжилттай нэмэгдлээ!', 'success')
        return redirect(url_for('vat'))
    partners = Partner.query.filter_by(is_active=True).order_by(Partner.name).all()
    return render_template('vat_form.html', today=date.today().isoformat(), partners=partners)

# ===================== CASH TRANSACTIONS =====================

@app.route('/cash')
def cash():
    from sqlalchemy import func as sqlfunc
    month        = request.args.get('month', type=int)
    bank         = request.args.get('bank', '')
    partner_q    = request.args.get('partner', '').strip()
    date_from    = request.args.get('date_from', '').strip()
    date_to      = request.args.get('date_to', '').strip()
    show_unlinked = request.args.get('unlinked', '')   # '1' = зөвхөн холбогдоогүй

    q = CashTransaction.query
    if month:
        q = q.filter_by(month=month)
    if bank:
        q = q.filter(CashTransaction.bank.like(f'%{bank}%'))
    if partner_q:
        matched_ids = [p.id for p in Partner.query.filter(
            _ilike(Partner.name, f'%{partner_q}%'), Partner.is_active == True
        ).all()]
        q = q.filter(db.or_(
            CashTransaction.partner_name.ilike(f'%{partner_q}%'),
            CashTransaction.master_name.ilike(f'%{partner_q}%'),
            CashTransaction.partner_id.in_(matched_ids) if matched_ids else db.false(),
        ))
    if date_from:
        try:
            q = q.filter(CashTransaction.date >= datetime.strptime(date_from, '%Y-%m-%d'))
        except ValueError:
            pass
    if date_to:
        try:
            q = q.filter(CashTransaction.date <= datetime.strptime(date_to + ' 23:59:59', '%Y-%m-%d %H:%M:%S'))
        except ValueError:
            pass
    if show_unlinked == '1':
        q = q.filter(db.or_(
            CashTransaction.debit_acc == None, CashTransaction.debit_acc == '',
            CashTransaction.credit_acc == None, CashTransaction.credit_acc == '',
        ))

    txns = q.order_by(CashTransaction.date.asc()).all()
    total_income   = sum(t.income  or 0 for t in txns)
    total_expense  = sum(t.expense or 0 for t in txns)
    unlinked_count = sum(1 for t in txns if not t.debit_acc or not t.credit_acc)

    banks = db.session.query(CashTransaction.bank).distinct().order_by(CashTransaction.bank).all()
    banks = [b[0] for b in banks if b[0]]
    months = db.session.query(CashTransaction.month).distinct().order_by(CashTransaction.month).all()
    months = [m[0] for m in months if m[0]]
    accounts = Account.query.filter_by(is_active=True).order_by(Account.code).all()
    partners_list = Partner.query.filter_by(is_active=True).order_by(Partner.name).all()

    return render_template('cash.html', txns=txns,
                           total_income=total_income, total_expense=total_expense,
                           unlinked_count=unlinked_count,
                           banks=banks, months=months,
                           sel_month=month, sel_bank=bank,
                           sel_partner=partner_q,
                           sel_date_from=date_from, sel_date_to=date_to,
                           sel_unlinked=show_unlinked,
                           accounts=accounts, partners_list=partners_list)

@app.route('/cash/auto-link', methods=['POST'])
def cash_auto_link():
    """Кодлох дүрмийн дагуу холбогдоогүй гүйлгээнүүдийг бөөнд холбоно"""
    txns = CashTransaction.query.filter(
        db.or_(
            CashTransaction.debit_acc == None,
            CashTransaction.debit_acc == '',
            CashTransaction.credit_acc == None,
            CashTransaction.credit_acc == ''
        ),
        db.or_(CashTransaction.income > 0, CashTransaction.expense > 0)
    ).all()

    linked = 0
    skipped = 0
    details = {}

    for t in txns:
        dt, kt = auto_link_cash_transaction(t)
        if dt and kt:
            t.debit_acc  = dt
            t.credit_acc = kt
            key = f'{t.income_cat or "–"}/{t.expense_cat or "–"} → {dt}/{kt}'
            details[key] = details.get(key, 0) + 1
            linked += 1
        else:
            skipped += 1

    db.session.commit()
    return jsonify({
        'linked':  linked,
        'skipped': skipped,
        'details': details,
        'message': f'{linked} гүйлгээ амжилттай холбогдлоо, {skipped} кодгүй алгасав'
    })

@app.route('/cash/summary')
def cash_summary():
    from sqlalchemy import func
    monthly = db.session.query(
        CashTransaction.month,
        func.sum(CashTransaction.income).label('income'),
        func.sum(CashTransaction.expense).label('expense')
    ).group_by(CashTransaction.month).order_by(CashTransaction.month).all()
    by_bank = db.session.query(
        CashTransaction.bank,
        func.sum(CashTransaction.income).label('income'),
        func.sum(CashTransaction.expense).label('expense')
    ).group_by(CashTransaction.bank).order_by(CashTransaction.bank).all()
    by_partner = db.session.query(
        CashTransaction.partner_name,
        func.sum(CashTransaction.income).label('income'),
        func.sum(CashTransaction.expense).label('expense')
    ).filter(CashTransaction.partner_name != None, CashTransaction.partner_name != '')\
     .group_by(CashTransaction.partner_name)\
     .order_by(func.sum(CashTransaction.income).desc()).limit(20).all()
    return render_template('cash_summary.html', monthly=monthly, by_bank=by_bank, by_partner=by_partner)


@app.route('/cash/bank-summary')
def cash_bank_summary():
    """Мөнгөн хөрөнгийн нэгтгэл — банк тус бүрийн үлдэгдэл, орлого, зарлага сараар"""
    from sqlalchemy import func, text

    BANKS = [
        ('1101', 'Голомт банк',  'Голомт'),
        ('1102', 'ХХБ / ТДБ',   'ХХБ'),
        ('1103', 'М банк',       'М банк'),
    ]
    ALL_MONTHS = list(range(1, 13))
    MONTH_NAMES = ['1-р', '2-р', '3-р', '4-р', '5-р', '6-р',
                   '7-р', '8-р', '9-р', '10-р', '11-р', '12-р']

    # ── 1. Эхний үлдэгдэл (journal_lines, posted, ≤ 2025-12-31) ──────────────
    #   DT 1101/1102/1103 дансаар + гарсан, KT дансаар - гарсан
    opening_raw = db.session.execute(text("""
        SELECT
            a.code,
            SUM(CASE WHEN jl.debit_account_id = a.id  THEN jl.amount ELSE 0 END) -
            SUM(CASE WHEN jl.credit_account_id = a.id THEN jl.amount ELSE 0 END) AS balance
        FROM accounts a
        JOIN journal_lines jl ON (jl.debit_account_id = a.id OR jl.credit_account_id = a.id)
        JOIN journals j ON j.id = jl.journal_id
        WHERE a.code IN ('1101','1102','1103')
          AND j.status = 'posted'
          AND j.date <= '2025-12-31'
        GROUP BY a.code
    """)).fetchall()
    opening_by_acc = {row[0]: float(row[1] or 0) for row in opening_raw}

    # ── 2. Банк тус бүрийн сарын орлого/зарлага cash_transactions-с ──────────
    # bank field нь текст тул _bank_code()-тэй ижил логик ашиглана
    inc_rows = db.session.query(
        CashTransaction.bank,
        CashTransaction.month,
        func.sum(CashTransaction.income).label('inc')
    ).filter(CashTransaction.income > 0)\
     .group_by(CashTransaction.bank, CashTransaction.month).all()

    exp_rows = db.session.query(
        CashTransaction.bank,
        CashTransaction.month,
        func.sum(CashTransaction.expense).label('exp')
    ).filter(CashTransaction.expense > 0)\
     .group_by(CashTransaction.bank, CashTransaction.month).all()

    def _classify(bank_str):
        if not bank_str: return '1101'
        b = str(bank_str)
        if '411096635' in b or 'ТДБ' in b or 'ХХБ' in b: return '1102'
        if '1175156757' in b or 'Голомт' in b:             return '1101'
        if '9006906192' in b or 'М банк' in b:             return '1103'
        return 'other'

    # inc_map[acc_code][month] = amount
    inc_map = {}
    for bank, mon, amt in inc_rows:
        code = _classify(bank)
        inc_map.setdefault(code, {})
        inc_map[code][mon or 0] = inc_map[code].get(mon or 0, 0) + float(amt or 0)

    exp_map = {}
    for bank, mon, amt in exp_rows:
        code = _classify(bank)
        exp_map.setdefault(code, {})
        exp_map[code][mon or 0] = exp_map[code].get(mon or 0, 0) + float(amt or 0)

    # ── 3. Сараар тооцоолол: эхний → орлого → зарлага → эцсийн ──────────────
    def build_bank(acc_code, open_bal):
        data = []
        bal = open_bal
        for m in ALL_MONTHS:
            inc = inc_map.get(acc_code, {}).get(m, 0)
            exp = exp_map.get(acc_code, {}).get(m, 0)
            net = inc - exp
            close = bal + net
            data.append({'m': m, 'open': bal, 'inc': inc, 'exp': exp, 'net': net, 'close': close})
            bal = close
        return data

    sections = []
    for acc_code, label, short in BANKS:
        open_bal = opening_by_acc.get(acc_code, 0)
        rows = build_bank(acc_code, open_bal)
        total_inc = sum(r['inc'] for r in rows)
        total_exp = sum(r['exp'] for r in rows)
        close_bal = rows[-1]['close'] if rows else open_bal
        sections.append({
            'code': acc_code, 'label': label, 'short': short,
            'open': open_bal, 'rows': rows,
            'total_inc': total_inc, 'total_exp': total_exp,
            'total_net': total_inc - total_exp,
            'close': close_bal,
        })

    # ── 4. Нэгтгэл (consolidated) ─────────────────────────────────────────────
    cons_open = sum(opening_by_acc.get(c, 0) for c, _, _ in BANKS)
    cons_rows = []
    bal = cons_open
    for i, m in enumerate(ALL_MONTHS):
        inc = sum(s['rows'][i]['inc'] for s in sections)
        exp = sum(s['rows'][i]['exp'] for s in sections)
        net = inc - exp
        close = bal + net
        cons_rows.append({'m': m, 'open': bal, 'inc': inc, 'exp': exp, 'net': net, 'close': close})
        bal = close
    cons_total_inc = sum(r['inc'] for r in cons_rows)
    cons_total_exp = sum(r['exp'] for r in cons_rows)
    consolidated = {
        'open': cons_open,
        'rows': cons_rows,
        'total_inc': cons_total_inc,
        'total_exp': cons_total_exp,
        'total_net': cons_total_inc - cons_total_exp,
        'close': cons_rows[-1]['close'] if cons_rows else cons_open,
    }

    return render_template('cash_bank_summary.html',
        sections=sections,
        consolidated=consolidated,
        months=MONTH_NAMES,
        year=2026,
    )


# ===================== PARTNERS =====================

@app.route('/api/partners/merge', methods=['POST'])
def partners_merge():
    """Сонгосон харилцагчдыг нэгтгэнэ — эхнийх нь үндсэн, бусад нь устана"""
    data       = request.get_json()
    ids        = data.get('ids', [])
    primary_id = data.get('primary_id')
    if len(ids) < 2:
        return jsonify({'error': 'Хамгийн багадаа 2 харилцагч сонгоно уу'}), 400

    primary = Partner.query.get(primary_id or ids[0])
    if not primary:
        return jsonify({'error': 'Үндсэн харилцагч олдсонгүй'}), 404

    others = [Partner.query.get(i) for i in ids if i != primary.id]
    others = [p for p in others if p]

    import json as _json
    try:
        aliases = _json.loads(primary.aliases) if primary.aliases else []
    except Exception:
        aliases = []

    merged_names = [primary.name]
    for dup in others:
        # Лавлагааг шилжүүлнэ
        VatRecord.query.filter_by(partner_id=dup.id).update(
            {'partner_id': primary.id}, synchronize_session=False)
        CashTransaction.query.filter_by(partner_id=dup.id).update(
            {'partner_id': primary.id}, synchronize_session=False)
        Receivable.query.filter_by(partner_id=dup.id).update(
            {'partner_id': primary.id}, synchronize_session=False)
        Payable.query.filter_by(partner_id=dup.id).update(
            {'partner_id': primary.id}, synchronize_session=False)

        # Нэрийг alias болгоно
        if dup.name and dup.name not in aliases and dup.name != primary.name:
            aliases.append(dup.name)
        merged_names.append(dup.name)

        # Мэдээлэл дутуу бол нөхнө
        if not primary.register and dup.register:
            primary.register = dup.register
        if not primary.code and dup.code:
            primary.code = dup.code
        if not primary.phone and dup.phone:
            primary.phone = dup.phone
        if not primary.email and dup.email:
            primary.email = dup.email

        dup.is_active = False

    primary.aliases = _json.dumps(aliases, ensure_ascii=False) if aliases else primary.aliases
    db.session.commit()

    return jsonify({
        'success': True,
        'primary': {'id': primary.id, 'name': primary.name},
        'merged': len(others),
        'message': f'{len(others)+1} харилцагч нэгтгэгдлээ → "{primary.name}"'
    })

@app.route('/api/partners/suggest')
def partners_suggest():
    """Харилцагчийн нэр autocomplete API"""
    q = request.args.get('q', '').strip()
    if len(q) < 1:
        return jsonify([])
    matches = Partner.query.filter(
        Partner.is_active == True,
        db.or_(
            _partner_search_conds(Partner.name, q, Partner.aliases),
            _ilike(Partner.code,     f'%{q}%'),
            _ilike(Partner.register, f'%{q}%'),
        )
    ).order_by(Partner.name).limit(15).all()
    return jsonify([{'id': p.id, 'name': p.name, 'code': p.code or '', 'register': p.register or ''} for p in matches])

@app.route('/partners')
def partners():
    from sqlalchemy import func as sqlfunc
    ptype = request.args.get('type', '')
    search = request.args.get('q', '').strip()

    q = Partner.query.filter_by(is_active=True)
    if ptype == 'customer':
        q = q.filter(db.or_(Partner.type == 'customer', Partner.type == 'both'))
    elif ptype == 'supplier':
        q = q.filter(db.or_(Partner.type == 'supplier', Partner.type == 'both'))
    if search:
        q = q.filter(db.or_(
            _partner_search_conds(Partner.name, search, Partner.aliases),
            _ilike(Partner.code,     f'%{search}%'),
            _ilike(Partner.register, f'%{search}%'),
        ))
    ps = q.order_by(Partner.name).all()

    # Банкны гүйлгээний нийт орлого/зарлага харилцагчаар
    cash_rows = db.session.query(
        CashTransaction.partner_id,
        sqlfunc.sum(CashTransaction.income).label('income'),
        sqlfunc.sum(CashTransaction.expense).label('expense'),
    ).filter(CashTransaction.partner_id != None)\
     .group_by(CashTransaction.partner_id).all()
    cash_by_pid = {r.partner_id: (r.income or 0, r.expense or 0) for r in cash_rows}

    # eBarimt нэхэмжлэлийн нийт дүн харилцагчаар
    vat_rows = db.session.query(
        VatRecord.partner_id,
        sqlfunc.sum(VatRecord.vat_amount * 11).label('total'),   # VAT 10% → total ≈ VAT*11
    ).filter(VatRecord.partner_id != None, VatRecord.type == 'out')\
     .group_by(VatRecord.partner_id).all()
    vat_by_pid = {r.partner_id: (r.total or 0) for r in vat_rows}

    total_income  = sum(v[0] for v in cash_by_pid.values())
    total_expense = sum(v[1] for v in cash_by_pid.values())

    return render_template('partners.html', partners=ps, sel_type=ptype,
                           search=search, cash_by_pid=cash_by_pid, vat_by_pid=vat_by_pid,
                           total_income=total_income, total_expense=total_expense)

@app.route('/partners/add', methods=['GET', 'POST'])
def add_partner():
    if request.method == 'POST':
        p = Partner(
            code=request.form.get('code', '').strip(),
            name=request.form.get('name', '').strip(),
            register=request.form.get('register', '').strip(),
            type=request.form.get('type', 'both'),
            phone=request.form.get('phone', ''),
            email=request.form.get('email', ''),
            address=request.form.get('address', '')
        )
        db.session.add(p)
        db.session.commit()
        flash(f'"{p.name}" харилцагч нэмэгдлээ!', 'success')
        return redirect(url_for('partners'))
    return render_template('partner_form.html', partner=None)

@app.route('/partners/edit/<int:id>', methods=['GET', 'POST'])
def edit_partner(id):
    p = Partner.query.get_or_404(id)
    if request.method == 'POST':
        p.code = request.form.get('code', '').strip()
        p.name = request.form.get('name', '').strip()
        p.register = request.form.get('register', '').strip()
        p.type = request.form.get('type', 'both')
        p.phone = request.form.get('phone', '')
        p.email = request.form.get('email', '')
        p.address = request.form.get('address', '')
        db.session.commit()
        flash('Харилцагч засагдлаа!', 'success')
        return redirect(url_for('partners'))
    return render_template('partner_form.html', partner=p)

@app.route('/partners/<int:id>')
def view_partner(id):
    p = Partner.query.get_or_404(id)

    # ── Огноо шүүлт ─────────────────────────────────────────────────────────
    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to',   '')
    try:
        df = datetime.strptime(date_from_str, '%Y-%m-%d').date()
    except Exception:
        df = None
    try:
        dt = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except Exception:
        dt = None

    receivables = Receivable.query.filter_by(partner_id=id).order_by(Receivable.date.desc()).all()
    payables    = Payable.query.filter_by(partner_id=id).order_by(Payable.date.desc()).all()

    # Банкны гүйлгээ — огноо шүүлттэй
    cash_q = CashTransaction.query.filter_by(partner_id=id)
    if df:
        cash_q = cash_q.filter(db.func.date(CashTransaction.date) >= df)
    if dt:
        cash_q = cash_q.filter(db.func.date(CashTransaction.date) <= dt)
    cash_txns = cash_q.order_by(CashTransaction.date.desc()).all()

    # E-Barimt — partner_id, регистр, нэр — OR-оор хайх
    vat_conds = [VatRecord.partner_id == id]
    if p.register:
        vat_conds.append(VatRecord.partner_register == p.register)
    if p.name:
        vat_conds.append(VatRecord.partner_name.ilike(p.name))
    vat_q = VatRecord.query.filter(db.or_(*vat_conds))
    if df:
        vat_q = vat_q.filter(VatRecord.date >= df)
    if dt:
        vat_q = vat_q.filter(VatRecord.date <= dt)
    # Давхардал арилгах (нэг record олон нөхцөлд таарсан бол)
    seen_vat_ids = set()
    vat_recs_raw = vat_q.order_by(VatRecord.date.desc()).all()
    vat_recs = []
    for v in vat_recs_raw:
        if v.id not in seen_vat_ids:
            seen_vat_ids.add(v.id)
            vat_recs.append(v)

    # Зөвхөн анхны eBarimt (parent_ddtd хоосон) = бодит борлуулалт
    # Хаалтын баримт (parent_ddtd бөглөгдсөн) = төлбөрийн баталгаа, орлого биш
    vat_out = [v for v in vat_recs if v.type == 'out' and not v.parent_ddtd]
    vat_out_closing = [v for v in vat_recs if v.type == 'out' and v.parent_ddtd]
    vat_in  = [v for v in vat_recs if v.type == 'in']

    total_recv         = sum(r.amount      for r in receivables)
    total_paid_recv    = sum(r.paid_amount for r in receivables)
    total_payable      = sum(r.amount      for r in payables)
    total_paid_payable = sum(r.paid_amount for r in payables)
    total_bank_income   = sum(t.income       for t in cash_txns)
    total_bank_expense  = sum(t.expense      for t in cash_txns)
    total_vat_out_total = sum(v.total_amount for v in vat_out)
    total_vat_in_total  = sum(v.total_amount for v in vat_in)

    return render_template('partner_view.html', partner=p,
                           receivables=receivables, payables=payables,
                           cash_txns=cash_txns,
                           total_recv=total_recv, total_paid_recv=total_paid_recv,
                           total_payable=total_payable, total_paid_payable=total_paid_payable,
                           vat_out=vat_out, vat_out_closing=vat_out_closing, vat_in=vat_in,
                           total_bank_income=total_bank_income,
                           total_bank_expense=total_bank_expense,
                           total_vat_out_total=total_vat_out_total,
                           total_vat_in_total=total_vat_in_total,
                           date_from=date_from_str, date_to=date_to_str)

# ===================== ТООЦОО НИЙЛСЭН АКТ (НХМаягт ТМ-3) =====================

@app.route('/partners/<int:id>/reconcile-act')
def reconcile_act(id):
    """НХМаягт ТМ-3 — Тооцооны үлдэгдлийн баталгаа"""
    p = Partner.query.get_or_404(id)

    date_from_str = request.args.get('date_from', '')
    date_to_str   = request.args.get('date_to',   '')
    act_no        = request.args.get('act_no', '')

    try:
        df = datetime.strptime(date_from_str, '%Y-%m-%d').date()
    except Exception:
        df = None
    try:
        dt = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except Exception:
        dt = None

    # ── Нэхэмжлэх (Receivable — Дебит) ─────────────────────────────────────
    recv_q = Receivable.query.filter_by(partner_id=id)
    if df: recv_q = recv_q.filter(Receivable.date >= df)
    if dt: recv_q = recv_q.filter(Receivable.date <= dt)
    recv_recs = recv_q.order_by(Receivable.date).all()

    # ── Банкны орлого (Кредит) ───────────────────────────────────────────────
    cash_q = CashTransaction.query.filter_by(partner_id=id).filter(CashTransaction.income > 0)
    if df: cash_q = cash_q.filter(db.func.date(CashTransaction.date) >= df)
    if dt: cash_q = cash_q.filter(db.func.date(CashTransaction.date) <= dt)
    cash_recs = cash_q.order_by(CashTransaction.date).all()

    # ── Мөрүүд нэгтгэж огноогоор эрэмбэлнэ ─────────────────────────────────
    rows = []
    for r in recv_recs:
        rows.append({
            'date':  r.date,
            'ref':   r.invoice_no or '',
            'desc':  r.description or f'Нэхэмжлэх — {p.name}',
            'debit': r.amount or 0,
            'credit': 0,
        })
    for c in cash_recs:
        rows.append({
            'date':   c.date.date() if hasattr(c.date, 'date') else c.date,
            'ref':    c.description or '',
            'desc':   c.description or f'Төлбөр — {p.name}',
            'debit':  0,
            'credit': c.income or 0,
        })
    rows.sort(key=lambda r: r['date'] or datetime.min.date())

    total_debit  = sum(r['debit']  for r in rows)
    total_credit = sum(r['credit'] for r in rows)
    net          = total_debit - total_credit   # + = авлага, - = өглөг
    ar_balance   = max(0,  net)
    ap_balance   = max(0, -net)

    return render_template('reconcile_act.html',
        partner=p, rows=rows,
        date_from=df, date_to=dt,
        act_no=act_no,
        total_debit=total_debit,
        total_credit=total_credit,
        net=net,
        ar_balance=ar_balance,
        ap_balance=ap_balance,
    )

# ===================== RECEIVABLES =====================

@app.route('/invoices')
def invoice_list():
    """Нэхэмжлэхийн жагсаалт — Receivable хүснэгтийн бүх бичлэг"""
    from sqlalchemy import func as sqlfunc
    status  = request.args.get('status', '')
    month   = request.args.get('month',  '', type=str)
    search  = request.args.get('q', '').strip()

    q = Receivable.query.join(Partner, Receivable.partner_id == Partner.id)
    if status:
        q = q.filter(Receivable.status == status)
    if month:
        try:
            m = int(month)
            q = q.filter(db.extract('month', Receivable.date) == m)
        except ValueError:
            pass
    if search:
        q = q.filter(db.or_(
            Receivable.invoice_no.ilike(f'%{search}%'),
            _ilike(Partner.name, f'%{search}%'),
            Receivable.description.ilike(f'%{search}%'),
        ))

    invs = q.order_by(Receivable.date.desc(), Receivable.invoice_no).all()

    total_amt      = sum(r.amount      for r in invs)
    total_paid     = sum(r.paid_amount for r in invs)
    total_remaining= sum(r.remaining   for r in invs)

    # Сарын нэгтгэл (шүүлтгүй нийт)
    monthly = db.session.query(
        db.extract('month', Receivable.date).label('m'),
        sqlfunc.count(Receivable.id).label('cnt'),
        sqlfunc.sum(Receivable.amount).label('amt'),
        sqlfunc.sum(Receivable.paid_amount).label('paid'),
    ).group_by('m').order_by('m').all()

    open_cnt    = Receivable.query.filter_by(status='open').count()
    partial_cnt = Receivable.query.filter_by(status='partial').count()
    paid_cnt    = Receivable.query.filter_by(status='paid').count()

    return render_template('invoice_list.html',
        invs=invs, total_amt=total_amt, total_paid=total_paid,
        total_remaining=total_remaining, sel_status=status,
        sel_month=month, search=search,
        monthly=monthly,
        open_cnt=open_cnt, partial_cnt=partial_cnt, paid_cnt=paid_cnt,
    )


@app.route('/invoices/<int:id>/print')
def invoice_print(id):
    """Нэхэмжлэхийн хэвлэх/PDF загвар"""
    inv = Receivable.query.get_or_404(id)
    # KAM (хариуцагч) ажилтныг нэрээр хайна
    kam = None
    if inv.responsible:
        term = inv.responsible.strip()
        kam = Employee.query.filter(
            db.or_(
                _ilike(Employee.name,     f'%{term}%'),
                _ilike(Employee.lastname, f'%{term}%'),
            ),
            Employee.is_active == True,
        ).first()
    return render_template('invoice_print.html', inv=inv, kam=kam)


@app.route('/receivables')
def receivables():
    status = request.args.get('status', '')

    # ── Харилцагчаар нэгтгэсэн авлага (E-Barimt + Банк) ──────────────────
    # Бүх идэвхтэй харилцагчийг уншина
    all_partners = Partner.query.filter_by(is_active=True).all()

    # VatRecord (type='out') → харилцагчаар нийт нэхэмжлэл
    # parent_ddtd-тэй баримтуудыг (дэд нэхэмжлэх) тоолохгүй — давхар тооцоогоос зайлсхийнэ
    from sqlalchemy import func as sqlfunc
    vat_rows = (db.session.query(VatRecord.partner_id, sqlfunc.sum(VatRecord.total_amount))
                .filter(VatRecord.type == 'out', VatRecord.partner_id != None,
                        db.or_(VatRecord.parent_ddtd == None, VatRecord.parent_ddtd == ''))
                .group_by(VatRecord.partner_id).all())
    vat_by_pid = {pid: amt for pid, amt in vat_rows}

    # VatRecord register-ээр (partner_id байхгүй үед)
    vat_reg_rows = (db.session.query(VatRecord.partner_register, sqlfunc.sum(VatRecord.total_amount))
                    .filter(VatRecord.type == 'out', VatRecord.partner_id == None,
                            VatRecord.partner_register != None,
                            db.or_(VatRecord.parent_ddtd == None, VatRecord.parent_ddtd == ''))
                    .group_by(VatRecord.partner_register).all())
    vat_by_reg = {reg: amt for reg, amt in vat_reg_rows}

    # CashTransaction income → харилцагчаар нийт орлого
    bank_rows = (db.session.query(CashTransaction.partner_id, sqlfunc.sum(CashTransaction.income))
                 .filter(CashTransaction.partner_id != None, CashTransaction.income > 0)
                 .group_by(CashTransaction.partner_id).all())
    bank_by_pid = {pid: amt for pid, amt in bank_rows}

    # Нэгтгэнэ
    recs = []
    seen_pids = set()
    for p in all_partners:
        invoiced = vat_by_pid.get(p.id, 0)
        if not invoiced and p.register:
            invoiced = vat_by_reg.get(p.register, 0)
        collected = bank_by_pid.get(p.id, 0)
        if invoiced == 0 and collected == 0:
            continue
        remaining = max(0, invoiced - collected)
        if invoiced > 0 and remaining < 1:
            st = 'paid'
        elif collected > 0 and remaining > 0:
            st = 'partial'
        elif invoiced > 0:
            st = 'open'
        else:
            st = 'open'
        if status and st != status:
            continue
        seen_pids.add(p.id)
        recs.append({
            'partner': p,
            'amount':    invoiced,
            'paid_amount': collected,
            'remaining': remaining,
            'status':    st,
        })

    # Мануал Receivable бичлэгүүд (хэрэв байвал нэмнэ)
    man_q = Receivable.query
    if status:
        man_q = man_q.filter_by(status=status)
    for r in man_q.order_by(Receivable.date.desc()).all():
        if r.partner_id not in seen_pids:
            recs.append({
                'partner':     r.partner,
                'amount':      r.amount,
                'paid_amount': r.paid_amount,
                'remaining':   r.remaining,
                'status':      r.status,
                'invoice_no':  r.invoice_no,
                'description': r.description,
            })

    recs.sort(key=lambda x: x['remaining'], reverse=True)
    total          = sum(r['amount']    for r in recs)
    total_collected = sum(r['paid_amount'] for r in recs)
    total_remaining = sum(r['remaining'] for r in recs)
    return render_template('receivables.html', receivables=recs, total=total,
                           total_collected=total_collected,
                           total_remaining=total_remaining, sel_status=status)

@app.route('/receivables/add', methods=['GET', 'POST'])
def add_receivable():
    if request.method == 'POST':
        date_val     = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        due_val      = datetime.strptime(request.form['due_date'], '%Y-%m-%d').date() if request.form.get('due_date') else None
        partner_id   = int(request.form['partner_id'])
        description  = request.form.get('description', '').strip()
        has_vat      = request.form.get('has_vat') == '1'
        base_amount  = float(request.form.get('base_amount', 0) or 0)
        vat_amount   = round(base_amount * 0.1, 2) if has_vat else 0
        total_amount = base_amount + vat_amount
        responsible  = request.form.get('responsible', '').strip()

        # Нэхэмжлэлийн дугаар автоматаар үүсгэнэ
        last = Receivable.query.order_by(Receivable.id.desc()).first()
        next_num = (last.id + 1) if last else 1
        invoice_no = request.form.get('invoice_no', '').strip() or f'INV-{date_val.year}-{next_num:04d}'

        r = Receivable(
            date=date_val,
            due_date=due_val,
            invoice_no=invoice_no,
            partner_id=partner_id,
            description=description,
            amount=total_amount,
            status='open',
            responsible=responsible,
        )
        db.session.add(r)
        db.session.flush()  # r.id авахын тулд

        # ── Тэмдэглэл: журналын бичилт хийхгүй ──────────────────────────────
        # Орлогын тооцоолол нь eBarimt журналаар (EB-OUT) хийгддэг.
        # Нэхэмжлэх бичлэг нь зөвхөн авлагын хяналтын баримт бичиг юм.
        # RECV-INV журнал бичвэл EB-OUT-тай давхардаж орлого хоёрдуулна.

        db.session.commit()
        flash(f'Нэхэмжлэх {invoice_no} амжилттай үүслээ!', 'success')
        # T-1 маягт хэвлэх хуудас руу шилжинэ
        return redirect(url_for('form_t1', rec=r.id))

    partners = Partner.query.filter_by(is_active=True).filter(
        db.or_(Partner.type == 'customer', Partner.type == 'both')).order_by(Partner.name).all()
    kam_employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    # Дараагийн нэхэмжлэлийн дугаар
    last = Receivable.query.order_by(Receivable.id.desc()).first()
    next_num = (last.id + 1) if last else 1
    next_inv  = f'INV-{date.today().year}-{next_num:04d}'
    return render_template('receivable_form.html', partners=partners, rec=None,
                           next_inv=next_inv, kam_employees=kam_employees)

@app.route('/receivables/<int:id>/edit', methods=['GET', 'POST'])
def edit_receivable(id):
    r = Receivable.query.get_or_404(id)
    if request.method == 'POST':
        r.responsible = request.form.get('responsible', '').strip()
        r.description = request.form.get('description', '').strip()
        due_str = request.form.get('due_date', '')
        r.due_date = datetime.strptime(due_str, '%Y-%m-%d').date() if due_str else None
        db.session.commit()
        flash(f'{r.invoice_no} — хариуцагч шинэчлэгдлээ.', 'success')
        return redirect(url_for('invoice_list'))

    kam_employees = Employee.query.filter_by(is_active=True).order_by(Employee.name).all()
    partners = Partner.query.filter_by(is_active=True).filter(
        db.or_(Partner.type == 'customer', Partner.type == 'both')).order_by(Partner.name).all()
    return render_template('receivable_form.html', rec=r, partners=partners,
                           next_inv=r.invoice_no, kam_employees=kam_employees)


@app.route('/receivables/<int:id>/pay', methods=['POST'])
def pay_receivable(id):
    r = Receivable.query.get_or_404(id)
    paid = float(request.form.get('amount', 0))
    r.paid_amount = min(r.amount, r.paid_amount + paid)
    if r.paid_amount >= r.amount:
        r.status = 'paid'
    elif r.paid_amount > 0:
        r.status = 'partial'
    db.session.commit()
    flash(f'{paid:,.0f}₮ төлөлт бүртгэгдлээ!', 'success')
    return redirect(url_for('receivables'))

# ===================== PAYABLES =====================

@app.route('/payables')
def payables():
    status         = request.args.get('status', '')
    partner_search = request.args.get('partner', '').strip().lower()
    from sqlalchemy import func as sqlfunc

    all_partners = Partner.query.filter_by(is_active=True).all()

    # ── VatRecord (type='in') → харилцагчаар нийт нэхэмжлэл ────────────────
    vat_rows = (db.session.query(VatRecord.partner_id, sqlfunc.sum(VatRecord.total_amount))
                .filter(VatRecord.type == 'in', VatRecord.partner_id != None)
                .group_by(VatRecord.partner_id).all())
    vat_by_pid = {pid: amt for pid, amt in vat_rows}

    vat_reg_rows = (db.session.query(VatRecord.partner_register, sqlfunc.sum(VatRecord.total_amount))
                    .filter(VatRecord.type == 'in', VatRecord.partner_id == None,
                            VatRecord.partner_register != None)
                    .group_by(VatRecord.partner_register).all())
    vat_by_reg = {reg: amt for reg, amt in vat_reg_rows}

    # ── Харилцагчаар дэлгэрэнгүй нэхэмжлэлүүд ──────────────────────────────
    vat_detail_rows = (VatRecord.query
                       .filter(VatRecord.type == 'in', VatRecord.partner_id != None)
                       .order_by(VatRecord.date.desc()).all())
    vat_detail_by_pid = {}
    for v in vat_detail_rows:
        vat_detail_by_pid.setdefault(v.partner_id, []).append(v)

    # ── CashTransaction expense → харилцагчаар нийт зарлага ─────────────────
    bank_rows = (db.session.query(CashTransaction.partner_id, sqlfunc.sum(CashTransaction.expense))
                 .filter(CashTransaction.partner_id != None, CashTransaction.expense > 0)
                 .group_by(CashTransaction.partner_id).all())
    bank_by_pid = {pid: amt for pid, amt in bank_rows}

    # ── Харилцагчаар дэлгэрэнгүй банкны зарлага ─────────────────────────────
    cash_detail_rows = (CashTransaction.query
                        .filter(CashTransaction.partner_id != None,
                                CashTransaction.expense > 0)
                        .order_by(CashTransaction.date.desc()).all())
    cash_detail_by_pid = {}
    for c in cash_detail_rows:
        cash_detail_by_pid.setdefault(c.partner_id, []).append(c)

    pays = []
    seen_pids = set()
    for p in all_partners:
        # Нэр хайлт шүүлт
        if partner_search and partner_search not in p.name.lower():
            continue
        invoiced = vat_by_pid.get(p.id, 0)
        if not invoiced and p.register:
            invoiced = vat_by_reg.get(p.register, 0)
        paid_amt = bank_by_pid.get(p.id, 0)
        if invoiced == 0 and paid_amt == 0:
            continue
        diff = invoiced - paid_amt
        remaining = max(0, diff)
        if invoiced > 0 and remaining < 1:
            st = 'paid'
        elif paid_amt > 0 and remaining > 0:
            st = 'partial'
        elif invoiced > 0:
            st = 'open'
        else:
            st = 'open'
        if status and st != status:
            continue
        seen_pids.add(p.id)
        pays.append({
            'partner':      p,
            'amount':       invoiced,
            'paid_amount':  paid_amt,
            'remaining':    remaining,
            'diff':         diff,        # can be negative (overpaid)
            'status':       st,
            'invoices':     vat_detail_by_pid.get(p.id, []),
            'payments':     cash_detail_by_pid.get(p.id, []),
        })

    pays.sort(key=lambda x: x['remaining'], reverse=True)
    total           = sum(r['amount']      for r in pays)
    total_paid      = sum(r['paid_amount'] for r in pays)
    total_remaining = sum(r['remaining']   for r in pays)
    return render_template('payables.html', payables=pays, total=total,
                           total_paid=total_paid, total_remaining=total_remaining,
                           sel_status=status, sel_partner=partner_search)

@app.route('/payables/<int:partner_id>/reconcile')
def reconcile_partner(partner_id):
    partner  = Partner.query.get_or_404(partner_id)
    invoices = VatRecord.query.filter_by(type='in', partner_id=partner_id)\
                              .order_by(VatRecord.date).all()
    payments = CashTransaction.query.filter(
        CashTransaction.partner_id == partner_id,
        CashTransaction.expense > 0
    ).order_by(CashTransaction.date).all()

    recons       = Reconciliation.query.filter_by(partner_id=partner_id).all()
    matched_inv  = {r.vat_record_id for r in recons if r.vat_record_id}
    matched_pay  = {r.cash_txn_id   for r in recons if r.cash_txn_id}

    unmatched_inv = [i for i in invoices if i.id not in matched_inv]
    unmatched_pay = [p for p in payments if p.id not in matched_pay]

    total_inv   = sum(i.total_amount or 0 for i in invoices)
    total_pay   = sum(p.expense or 0 for p in payments)
    total_recon = sum(r.matched_amount for r in recons)

    return render_template('reconcile.html',
        partner=partner, invoices=invoices, payments=payments,
        recons=recons, unmatched_inv=unmatched_inv, unmatched_pay=unmatched_pay,
        total_inv=total_inv, total_pay=total_pay, total_recon=total_recon)

@app.route('/api/reconcile/auto-match', methods=['POST'])
def api_reconcile_auto():
    """Автоматаар яг тохирч байгаа дүнгүүдийг тулгана"""
    partner_id = request.json.get('partner_id')
    recons = Reconciliation.query.filter_by(partner_id=partner_id).all()
    matched_inv = {r.vat_record_id for r in recons if r.vat_record_id}
    matched_pay = {r.cash_txn_id   for r in recons if r.cash_txn_id}

    invoices = VatRecord.query.filter_by(type='in', partner_id=partner_id).all()
    payments = CashTransaction.query.filter(
        CashTransaction.partner_id == partner_id, CashTransaction.expense > 0
    ).all()

    unmatched_inv = [i for i in invoices if i.id not in matched_inv]
    unmatched_pay = [p for p in payments if p.id not in matched_pay]

    # amount → list of payments (same amount can repeat)
    pay_map = {}
    for p in unmatched_pay:
        key = round(p.expense or 0)
        pay_map.setdefault(key, []).append(p)

    created = 0
    for inv in unmatched_inv:
        key = round(inv.total_amount or 0)
        if pay_map.get(key):
            pay = pay_map[key].pop(0)
            r = Reconciliation(partner_id=partner_id,
                               vat_record_id=inv.id, cash_txn_id=pay.id,
                               matched_amount=inv.total_amount, diff=0)
            db.session.add(r)
            created += 1
    db.session.commit()
    return jsonify({'ok': True, 'created': created})

@app.route('/api/reconcile/match', methods=['POST'])
def api_reconcile_match():
    """Нэхэмжлэл болон төлбөрийг гараар тулгана"""
    data       = request.json
    partner_id = data.get('partner_id')
    inv_id     = data.get('invoice_id')
    pay_id     = data.get('payment_id')
    notes      = data.get('notes', '')

    inv = VatRecord.query.get(inv_id)       if inv_id else None
    pay = CashTransaction.query.get(pay_id) if pay_id else None

    inv_amt = inv.total_amount if inv else 0
    pay_amt = pay.expense      if pay else 0
    matched = min(inv_amt, pay_amt)
    diff    = inv_amt - pay_amt

    # Хэрэв зөрүү байвал journal entry үүсгэх эсэхийг шалгах
    create_je = data.get('create_journal', False)
    if create_je and abs(diff) > 0.5:
        diff_acc_code = '8312' if diff > 0 else '1250'   # банкны шимтгэл / бусад авлага
        diff_acc = Account.query.filter_by(code=diff_acc_code, is_active=True).first()
        pay_acc  = Account.query.filter_by(code='3310',  is_active=True).first()
        if diff_acc and pay_acc:
            j = Journal(date=date.today(), number=get_next_journal_number(),
                description=f'Тулгалтын зөрүү — {inv.partner_name if inv else ""}',
                reference='RECON', status='posted')
            db.session.add(j); db.session.flush()
            if diff > 0:
                db.session.add(JournalLine(journal_id=j.id,
                    debit_account_id=diff_acc.id, credit_account_id=pay_acc.id,
                    amount=abs(diff), description='Тулгалтын зөрүү (хасагдуулга)'))
            else:
                db.session.add(JournalLine(journal_id=j.id,
                    debit_account_id=pay_acc.id, credit_account_id=diff_acc.id,
                    amount=abs(diff), description='Тулгалтын зөрүү (буцаалт)'))

    r = Reconciliation(partner_id=partner_id, vat_record_id=inv_id, cash_txn_id=pay_id,
                       matched_amount=matched, diff=diff, notes=notes)
    db.session.add(r)
    db.session.commit()
    return jsonify({'ok': True, 'id': r.id, 'matched': matched, 'diff': diff})

@app.route('/api/reconcile/<int:rid>', methods=['DELETE'])
def api_reconcile_delete(rid):
    r = Reconciliation.query.get_or_404(rid)
    db.session.delete(r)
    db.session.commit()
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════════════════════
# НЭХЭМЖЛЭЛ ↔ БАНК ТУЛГАЛТ  (InvoicePayment)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/recon/invoices')
def invoice_bank_recon():
    from sqlalchemy import func as sqlfunc
    sel_partner = request.args.get('partner_id', '', type=str)
    sel_month   = request.args.get('month',      '', type=str)
    sel_status  = request.args.get('status',     '', type=str)

    # Хүснэгт: харилцагчаар нэхэмжлэл нийт
    inv_q = db.session.query(
        Receivable.partner_id,
        sqlfunc.sum(Receivable.amount).label('total_inv'),
        sqlfunc.sum(Receivable.paid_amount).label('paid_inv'),
        sqlfunc.count(Receivable.id).label('cnt_inv'),
    ).group_by(Receivable.partner_id)
    if sel_month:
        try:
            inv_q = inv_q.filter(db.extract('month', Receivable.date) == int(sel_month))
        except ValueError:
            pass
    inv_rows = inv_q.all()
    inv_by_pid = {r.partner_id: r for r in inv_rows}

    # Хүснэгт: харилцагчаар банкны орлого нийт
    bank_q = db.session.query(
        CashTransaction.partner_id,
        sqlfunc.sum(CashTransaction.income).label('total_bank'),
        sqlfunc.count(CashTransaction.id).label('cnt_bank'),
    ).filter(CashTransaction.partner_id != None, CashTransaction.income > 0)\
     .group_by(CashTransaction.partner_id)
    if sel_month:
        try:
            bank_q = bank_q.filter(CashTransaction.month == int(sel_month))
        except ValueError:
            pass
    bank_rows = bank_q.all()
    bank_by_pid = {r.partner_id: r for r in bank_rows}

    # Тулгасан дүн харилцагчаар
    match_rows = db.session.query(
        Receivable.partner_id,
        sqlfunc.sum(InvoicePayment.amount).label('matched'),
    ).join(InvoicePayment, InvoicePayment.receivable_id == Receivable.id)\
     .group_by(Receivable.partner_id).all()
    matched_by_pid = {r.partner_id: (r.matched or 0) for r in match_rows}

    all_pids = set(inv_by_pid.keys()) | set(bank_by_pid.keys())
    partners_map = {p.id: p for p in Partner.query.filter(Partner.id.in_(all_pids)).all()}

    rows = []
    for pid in all_pids:
        if sel_partner and str(pid) != sel_partner:
            continue
        p = partners_map.get(pid)
        if not p:
            continue
        inv  = inv_by_pid.get(pid)
        bank = bank_by_pid.get(pid)
        total_inv   = (inv.total_inv   if inv  else 0) or 0
        total_bank  = (bank.total_bank if bank else 0) or 0
        matched     = matched_by_pid.get(pid, 0)
        cnt_inv     = (inv.cnt_inv     if inv  else 0) or 0
        cnt_bank    = (bank.cnt_bank   if bank else 0) or 0
        diff        = total_bank - total_inv
        if abs(diff) < 1:
            status = 'match'
        elif diff > 0:
            status = 'bank_more'
        else:
            status = 'inv_more'
        rows.append({
            'partner_id':   pid,
            'partner_name': p.name,
            'total_inv':    total_inv,
            'total_bank':   total_bank,
            'matched':      matched,
            'cnt_inv':      cnt_inv,
            'cnt_bank':     cnt_bank,
            'diff':         diff,
            'status':       status,
        })

    rows.sort(key=lambda x: -abs(x['diff']))

    # Статусаар тоо ба дүн — шүүлтийн ӨМНӨ тооцоолно (бүх rows-аас)
    cnt_total    = len(rows)
    cnt_match    = sum(1 for r in rows if r['status'] == 'match')
    cnt_inv_more = sum(1 for r in rows if r['status'] == 'inv_more')
    cnt_bank_more= sum(1 for r in rows if r['status'] == 'bank_more')

    # Stat card-д харуулах дүнгүүд (бүх rows-аас)
    amt_inv_all   = sum(r['total_inv']  for r in rows)
    amt_bank_all  = sum(r['total_bank'] for r in rows)
    amt_match_all = sum(r['matched']    for r in rows)
    amt_inv_more  = sum(abs(r['diff'])  for r in rows if r['status'] == 'inv_more')
    amt_bank_more = sum(r['diff']       for r in rows if r['status'] == 'bank_more')

    # Хэрэв status шүүлт байвал rows-г шүүнэ
    if sel_status:
        rows = [r for r in rows if r['status'] == sel_status]

    total_inv_s   = sum(r['total_inv']  for r in rows)
    total_bank_s  = sum(r['total_bank'] for r in rows)
    total_match_s = sum(r['matched']    for r in rows)

    partners_list = [partners_map[pid] for pid in sorted(all_pids) if pid in partners_map]

    return render_template('invoice_bank_recon.html',
        rows=rows,
        partners_list=partners_list,
        sel_partner=sel_partner,
        sel_month=sel_month,
        sel_status=sel_status,
        total_inv=total_inv_s,
        total_bank=total_bank_s,
        total_match=total_match_s,
        cnt_total=cnt_total,
        cnt_match=cnt_match,
        cnt_inv_more=cnt_inv_more,
        cnt_bank_more=cnt_bank_more,
        amt_inv_more=amt_inv_more,
        amt_bank_more=amt_bank_more,
        # stat card-д бүх rows-ийн дүн хэрэгтэй
        stat_inv=amt_inv_all,
        stat_match=amt_match_all,
    )


@app.route('/api/inv-recon/partner-detail')
def api_inv_recon_detail():
    from sqlalchemy import func as sqlfunc
    pid = request.args.get('partner_id', type=int)
    if not pid:
        return jsonify({'error': 'No partner'}), 400

    invs = Receivable.query.filter_by(partner_id=pid).order_by(Receivable.date).all()
    inv_data = []
    for r in invs:
        m = db.session.query(sqlfunc.sum(InvoicePayment.amount))\
              .filter_by(receivable_id=r.id).scalar() or 0
        pm_list = InvoicePayment.query.filter_by(receivable_id=r.id).all()
        inv_data.append({
            'id':         r.id,
            'invoice_no': r.invoice_no or '',
            'date':       str(r.date),
            'due_date':   str(r.due_date) if r.due_date else '',
            'amount':     r.amount,
            'paid':       r.paid_amount,
            'matched':    m,
            'remaining':  r.amount - r.paid_amount,
            'status':     r.status,
            'description':r.description or '',
            'payments':   [{'id': p.id, 'cash_txn_id': p.cash_txn_id,
                            'amount': p.amount} for p in pm_list],
        })

    txns = CashTransaction.query.filter_by(partner_id=pid)\
             .filter(CashTransaction.income > 0)\
             .order_by(CashTransaction.date).all()
    txn_data = []
    for t in txns:
        m = db.session.query(sqlfunc.sum(InvoicePayment.amount))\
              .filter_by(cash_txn_id=t.id).scalar() or 0
        txn_data.append({
            'id':          t.id,
            'date':        str(t.date)[:10],
            'description': t.description or '',
            'income':      t.income or 0,
            'matched':     m,
            'unmatched':   (t.income or 0) - m,
            'bank':        t.bank or '',
        })

    return jsonify({'invoices': inv_data, 'transactions': txn_data})


@app.route('/api/inv-recon/match', methods=['POST'])
def api_inv_recon_match():
    from sqlalchemy import func as sqlfunc
    data    = request.get_json() or {}
    recv_id = data.get('receivable_id')
    txn_id  = data.get('cash_txn_id')
    amount  = float(data.get('amount', 0))

    r = Receivable.query.get_or_404(recv_id)
    t = CashTransaction.query.get_or_404(txn_id)

    inv_remaining = r.amount - r.paid_amount
    txn_matched   = db.session.query(sqlfunc.sum(InvoicePayment.amount))\
                      .filter_by(cash_txn_id=txn_id).scalar() or 0
    txn_unmatched = (t.income or 0) - txn_matched

    if amount <= 0:
        amount = min(inv_remaining, txn_unmatched)
    if amount <= 0.01:
        return jsonify({'ok': False, 'error': 'Тулгах дүн байхгүй'}), 400

    pm = InvoicePayment(receivable_id=recv_id, cash_txn_id=txn_id, amount=amount)
    db.session.add(pm)

    r.paid_amount = min(r.amount, r.paid_amount + amount)
    r.status = 'paid' if r.paid_amount >= r.amount else ('partial' if r.paid_amount > 0 else 'open')
    db.session.commit()
    return jsonify({'ok': True, 'id': pm.id, 'amount': amount,
                    'inv_paid': r.paid_amount, 'inv_status': r.status})


@app.route('/api/inv-recon/unmatch/<int:pm_id>', methods=['DELETE'])
def api_inv_recon_unmatch(pm_id):
    pm = InvoicePayment.query.get_or_404(pm_id)
    r  = pm.receivable
    r.paid_amount = max(0, r.paid_amount - pm.amount)
    r.status = 'paid' if r.paid_amount >= r.amount else ('partial' if r.paid_amount > 0 else 'open')
    db.session.delete(pm)
    db.session.commit()
    return jsonify({'ok': True})


@app.route('/api/inv-recon/auto-match', methods=['POST'])
def api_inv_recon_auto():
    from sqlalchemy import func as sqlfunc
    data = request.get_json() or {}
    pid  = data.get('partner_id')
    if not pid:
        return jsonify({'ok': False, 'error': 'partner_id required'}), 400

    invs = Receivable.query.filter_by(partner_id=pid)\
             .filter(Receivable.status.in_(['open', 'partial']))\
             .order_by(Receivable.date).all()
    txns = CashTransaction.query.filter_by(partner_id=pid)\
             .filter(CashTransaction.income > 0)\
             .order_by(CashTransaction.date).all()

    matched_count = 0
    for t in txns:
        already = db.session.query(sqlfunc.sum(InvoicePayment.amount))\
                    .filter_by(cash_txn_id=t.id).scalar() or 0
        rem_bank = (t.income or 0) - already
        for r in invs:
            if rem_bank <= 0.01:
                break
            inv_rem = r.amount - r.paid_amount
            if inv_rem <= 0.01:
                continue
            match_amt = min(rem_bank, inv_rem)
            db.session.add(InvoicePayment(receivable_id=r.id, cash_txn_id=t.id, amount=match_amt))
            r.paid_amount = min(r.amount, r.paid_amount + match_amt)
            r.status = 'paid' if r.paid_amount >= r.amount else 'partial'
            rem_bank -= match_amt
            matched_count += 1

    db.session.commit()
    return jsonify({'ok': True, 'matched': matched_count})


@app.route('/payables/add', methods=['GET', 'POST'])
def add_payable():
    if request.method == 'POST':
        p = Payable(
            date=datetime.strptime(request.form['date'], '%Y-%m-%d').date(),
            due_date=datetime.strptime(request.form['due_date'], '%Y-%m-%d').date() if request.form.get('due_date') else None,
            invoice_no=request.form.get('invoice_no', ''),
            partner_id=int(request.form['partner_id']),
            description=request.form.get('description', ''),
            amount=float(request.form.get('amount', 0)),
            status='open'
        )
        db.session.add(p)
        db.session.commit()
        flash('Өглөг нэмэгдлээ!', 'success')
        return redirect(url_for('payables'))
    partners = Partner.query.filter_by(is_active=True).filter(
        db.or_(Partner.type == 'supplier', Partner.type == 'both')).order_by(Partner.name).all()
    return render_template('payable_form.html', partners=partners, pay=None)

@app.route('/payables/<int:id>/pay', methods=['POST'])
def pay_payable(id):
    p = Payable.query.get_or_404(id)
    paid = float(request.form.get('amount', 0))
    p.paid_amount = min(p.amount, p.paid_amount + paid)
    if p.paid_amount >= p.amount:
        p.status = 'paid'
    elif p.paid_amount > 0:
        p.status = 'partial'
    db.session.commit()
    flash(f'{paid:,.0f}₮ төлөлт бүртгэгдлээ!', 'success')
    return redirect(url_for('payables'))

# ---- Банкны гүйлгээний багцын журнал ----
@app.route('/cash/batch-journal', methods=['GET', 'POST'])
def cash_batch_journal():
    from sqlalchemy import func as sqlfunc

    acc_lookup = {a.code: a for a in Account.query.filter_by(is_active=True).all()}

    # Сарын preview — (year, month) бүрийн орлого/зарлага, хос тоо, бичигдсэн эсэх
    monthly_raw = db.session.query(
        CashTransaction.year,
        CashTransaction.month,
        sqlfunc.count().label('cnt'),
        sqlfunc.sum(CashTransaction.income).label('inc'),
        sqlfunc.sum(CashTransaction.expense).label('exp'),
        sqlfunc.count(CashTransaction.journal_id).label('done_cnt'),
    ).group_by(CashTransaction.year, CashTransaction.month)\
     .order_by(CashTransaction.year, CashTransaction.month).all()

    preview = []
    for r in monthly_raw:
        num = f"BNK-{r.year}-{r.month:02d}"
        j = Journal.query.filter_by(number=num).first()
        preview.append({
            'year': r.year, 'month': r.month, 'num': num,
            'cnt': r.cnt,
            'inc': r.inc or 0,
            'exp': r.exp or 0,
            'done': j is not None,
            'journal_id': j.id if j else None,
        })

    if request.method == 'POST':
        selected = request.form.getlist('months')  # "YYYY-MM" list
        created = 0
        skipped = 0
        errors = []

        for ym in selected:
            try:
                yr, mo = int(ym.split('-')[0]), int(ym.split('-')[1])
            except:
                continue

            num = f"BNK-{yr}-{mo:02d}"
            if Journal.query.filter_by(number=num).first():
                skipped += 1
                continue

            # Тухайн сарын бүх гүйлгээг (dt, kt) хосоор нэгтгэнэ
            pairs = db.session.query(
                CashTransaction.debit_acc,
                CashTransaction.credit_acc,
                sqlfunc.sum(CashTransaction.income + CashTransaction.expense).label('total'),
            ).filter_by(year=yr, month=mo)\
             .group_by(CashTransaction.debit_acc, CashTransaction.credit_acc).all()

            lines = []
            for p in pairs:
                dt_code = str(p.debit_acc or '').strip()
                kt_code = str(p.credit_acc or '').strip()
                amt     = round(p.total or 0, 2)
                if amt <= 0 or not dt_code or not kt_code:
                    continue
                dt_acc = acc_lookup.get(dt_code)
                kt_acc = acc_lookup.get(kt_code)
                if not dt_acc or not kt_acc:
                    errors.append(f"{num}: данс олдсонгүй DT={dt_code} KT={kt_code}")
                    continue
                lines.append(JournalLine(
                    debit_account_id=dt_acc.id,
                    credit_account_id=kt_acc.id,
                    amount=amt,
                    description=f"Банкны хуулга {yr}/{mo:02d}",
                ))

            if not lines:
                errors.append(f"{num}: бичилтийн мөр алга")
                continue

            # Сарын эхний гүйлгээний огноог ашиглана
            first_txn = CashTransaction.query.filter_by(year=yr, month=mo)\
                            .order_by(CashTransaction.date).first()
            jdate = first_txn.date.date() if first_txn and first_txn.date else date.today()

            j = Journal(
                number=num,
                date=jdate,
                description=f"Банкны хуулга {yr}/{mo:02d}",
                status='posted',
            )
            j.lines = lines
            db.session.add(j)
            db.session.flush()  # j.id авна

            # CashTransaction-д journal_id тохируулна
            CashTransaction.query.filter_by(year=yr, month=mo)\
                .update({'journal_id': j.id})

            created += 1

        db.session.commit()

        msg = f"{created} сарын журнал үүсгэгдлээ"
        if skipped: msg += f", {skipped} давхардал алгасав"
        if errors:  msg += f". Анхаар: {'; '.join(errors[:3])}"
        flash(msg, 'success' if not errors else 'warning')
        return redirect(url_for('cash_batch_journal'))

    return render_template('cash_batch_journal.html', preview=preview)


# ---- Мөнгөн гүйлгээний журнаалын бичилт ----
@app.route('/cash/<int:id>/journal')
def cash_journal(id):
    t = CashTransaction.query.get_or_404(id)
    if not t.journal_id:
        return jsonify({'ok': True, 'lines': [], 'journal_no': None})
    j = Journal.query.get(t.journal_id)
    if not j:
        return jsonify({'ok': True, 'lines': [], 'journal_no': None})
    lines = []
    for l in j.lines:
        dt_acc = Account.query.get(l.debit_account_id)  if l.debit_account_id  else None
        kt_acc = Account.query.get(l.credit_account_id) if l.credit_account_id else None
        lines.append({
            'dt_code': dt_acc.code if dt_acc else '—',
            'dt_name': dt_acc.name if dt_acc else '—',
            'kt_code': kt_acc.code if kt_acc else '—',
            'kt_name': kt_acc.name if kt_acc else '—',
            'amount':  l.amount,
            'desc':    l.description or '',
        })
    return jsonify({'ok': True, 'lines': lines,
                    'journal_no': j.number, 'journal_date': str(j.date),
                    'journal_desc': j.description or ''})

# ---- Мөнгөн гүйлгээний ангилал засах ----
@app.route('/cash/<int:id>/update-cat', methods=['POST'])
def update_cash_cat(id):
    t = CashTransaction.query.get_or_404(id)
    if t.income:
        t.income_cat = request.json.get('cat', '').strip() or t.income_cat
    else:
        t.expense_cat = request.json.get('cat', '').strip() or t.expense_cat
    db.session.commit()
    return jsonify({'ok': True, 'income_cat': t.income_cat, 'expense_cat': t.expense_cat})

# ---- Мөнгөн гүйлгээний ДТ/КТ данс засах ----
@app.route('/cash/<int:id>/update-accounts', methods=['POST'])
def update_cash_accounts(id):
    t = CashTransaction.query.get_or_404(id)
    new_dt = request.form.get('debit_acc', '').strip()
    new_kt = request.form.get('credit_acc', '').strip()
    t.debit_acc  = new_dt or t.debit_acc
    t.credit_acc = new_kt or t.credit_acc
    db.session.commit()

    # Sync the linked CASH journal if it exists
    jnl = Journal.query.filter_by(number=f'CASH-{id:05d}').first()
    if jnl and jnl.lines:
        dt_acc = Account.query.filter_by(code=t.debit_acc, is_active=True).first()
        kt_acc = Account.query.filter_by(code=t.credit_acc, is_active=True).first()
        if dt_acc and kt_acc:
            line = jnl.lines[0]
            line.debit_account_id  = dt_acc.id
            line.credit_account_id = kt_acc.id
            db.session.commit()

    return jsonify({'ok': True, 'dt': t.debit_acc, 'kt': t.credit_acc})

# ===================== ҮНДСЭН ХӨРӨНГӨ =====================

ASSET_CATEGORIES = [
    'Тоног төхөөрөмж', 'Тээврийн хэрэгсэл', 'Барилга, байгууламж',
    'Тавилга, эд хогшил', 'Компьютер, техник', 'Биет бус хөрөнгө', 'Бусад'
]

@app.route('/fixed-assets')
def fixed_assets():
    cat    = request.args.get('cat', '')
    status = request.args.get('status', 'active')
    q = FixedAsset.query
    if cat:
        q = q.filter_by(category=cat)
    if status:
        q = q.filter_by(status=status)
    assets = q.order_by(FixedAsset.code).all()
    total_cost  = sum(a.purchase_amount for a in assets)
    total_depr  = sum(a.accumulated_depreciation for a in assets)
    total_book  = sum(a.book_value for a in assets)
    return render_template('fixed_assets.html', assets=assets,
                           total_cost=total_cost, total_depr=total_depr, total_book=total_book,
                           categories=ASSET_CATEGORIES, sel_cat=cat, sel_status=status)

@app.route('/fixed-assets/add', methods=['GET', 'POST'])
def add_fixed_asset():
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        if code and FixedAsset.query.filter_by(code=code).first():
            flash('Энэ кодтой хөрөнгө аль хэдийн байна!', 'danger')
            return redirect(url_for('add_fixed_asset'))
        pd_str = request.form.get('purchase_date', '')
        fa = FixedAsset(
            code=code or None,
            name=request.form.get('name', '').strip(),
            category=request.form.get('category', 'Тоног төхөөрөмж'),
            purchase_date=datetime.strptime(pd_str, '%Y-%m-%d').date() if pd_str else None,
            purchase_amount=float(request.form.get('purchase_amount', 0)),
            useful_life=int(request.form.get('useful_life', 5)),
            salvage_value=float(request.form.get('salvage_value', 0)),
            accumulated_depreciation=float(request.form.get('accumulated_depreciation', 0)),
            location=request.form.get('location', ''),
            responsible_person=request.form.get('responsible_person', ''),
            status=request.form.get('status', 'active'),
            notes=request.form.get('notes', '')
        )
        db.session.add(fa)
        db.session.commit()
        flash(f'"{fa.name}" хөрөнгө нэмэгдлээ!', 'success')
        return redirect(url_for('fixed_assets'))
    return render_template('fixed_asset_form.html', asset=None, categories=ASSET_CATEGORIES,
                           today=date.today().isoformat())

@app.route('/fixed-assets/edit/<int:id>', methods=['GET', 'POST'])
def edit_fixed_asset(id):
    fa = FixedAsset.query.get_or_404(id)
    if request.method == 'POST':
        pd_str = request.form.get('purchase_date', '')
        fa.code         = request.form.get('code', '').strip() or fa.code
        fa.name         = request.form.get('name', fa.name).strip()
        fa.category     = request.form.get('category', fa.category)
        fa.purchase_date= datetime.strptime(pd_str, '%Y-%m-%d').date() if pd_str else fa.purchase_date
        fa.purchase_amount = float(request.form.get('purchase_amount', fa.purchase_amount))
        fa.useful_life  = int(request.form.get('useful_life', fa.useful_life))
        fa.salvage_value= float(request.form.get('salvage_value', fa.salvage_value))
        fa.accumulated_depreciation = float(request.form.get('accumulated_depreciation', fa.accumulated_depreciation))
        fa.location     = request.form.get('location', fa.location)
        fa.responsible_person = request.form.get('responsible_person', fa.responsible_person)
        fa.status       = request.form.get('status', fa.status)
        fa.notes        = request.form.get('notes', fa.notes)
        db.session.commit()
        flash('Хөрөнгийн мэдээлэл засагдлаа!', 'success')
        return redirect(url_for('view_fixed_asset', id=fa.id))
    return render_template('fixed_asset_form.html', asset=fa, categories=ASSET_CATEGORIES,
                           today=date.today().isoformat())

@app.route('/fixed-assets/<int:id>')
def view_fixed_asset(id):
    fa = FixedAsset.query.get_or_404(id)
    # Monthly depreciation schedule
    schedule = []
    if fa.purchase_date and fa.useful_life and fa.annual_depreciation > 0:
        monthly = fa.monthly_depreciation
        accum   = 0
        from datetime import timedelta
        cur_date = fa.purchase_date.replace(day=1)
        months_total = fa.useful_life * 12
        for i in range(min(months_total, 120)):   # max 10 yrs shown
            accum = min(accum + monthly, fa.purchase_amount - fa.salvage_value)
            schedule.append({
                'month': cur_date.strftime('%Y-%m'),
                'amount': monthly,
                'accum': accum,
                'book': fa.purchase_amount - accum
            })
            # next month
            if cur_date.month == 12:
                cur_date = cur_date.replace(year=cur_date.year+1, month=1)
            else:
                cur_date = cur_date.replace(month=cur_date.month+1)
    return render_template('fixed_asset_view.html', asset=fa, schedule=schedule)

@app.route('/fixed-assets/<int:id>/depreciate', methods=['POST'])
def depreciate_fixed_asset(id):
    fa = FixedAsset.query.get_or_404(id)
    # Accept either explicit amount or months
    amount_raw = request.form.get('amount', '')
    if amount_raw:
        add_depr = float(amount_raw)
    else:
        months = int(request.form.get('months', 1))
        add_depr = fa.monthly_depreciation * months

    max_depr = fa.purchase_amount - fa.salvage_value
    add_depr = min(add_depr, max(0, max_depr - fa.accumulated_depreciation))
    if add_depr <= 0:
        flash('Элэгдэл бичих дүн тэг эсвэл хөрөнгө бүрэн элэгдсэн байна!', 'danger')
        return redirect(url_for('fixed_assets'))

    fa.accumulated_depreciation += add_depr

    # ─── Journal entry: DT=5300 (Элэгдлийн зардал) / KT=201xxx (Хуримтлагдсан элэгдэл) ───
    # Map category → accumulated-depreciation account code
    _CAT_DEPR_ACC = {
        'Барилга, байгуулам':          '201201',
        'Машин, тоног төхөөрөмж':      '201301',
        'Тээврийн хэрэгсэл':           '201401',
        'Тавилга эд хогшил':           '201501',
        'Компьютер, бусад хэрэгсэл':   '201601',
        'Бусад үндсэн хөрөнгө':        '201701',
        'Холбооны тоног төхөөрөмж':    '201801',
    }
    depr_exp_acc  = Account.query.filter_by(code='5300').first()
    depr_depr_code = _CAT_DEPR_ACC.get(fa.category, '201701')
    depr_acc_acc  = Account.query.filter(Account.code == depr_depr_code).first()

    if depr_exp_acc and depr_acc_acc:
        j = Journal(
            date=date.today(),
            number=get_next_journal_number(),
            description=f'Элэгдэл — {fa.name}',
            reference=fa.code or '',
            status='posted'
        )
        db.session.add(j)
        db.session.flush()
        db.session.add(JournalLine(
            journal_id=j.id,
            debit_account_id=depr_exp_acc.id,
            credit_account_id=depr_acc_acc.id,
            amount=add_depr,
            description=f'{fa.name} элэгдэл'
        ))
        flash(f'Элэгдэл {add_depr:,.0f}₮ бүртгэгдлээ! Журнаал: {j.number}', 'success')
    else:
        flash(f'Элэгдэл {add_depr:,.0f}₮ бүртгэгдлээ! (5300 эсвэл {depr_depr_code} данс олдсонгүй)', 'warning')

    db.session.commit()
    return redirect(url_for('fixed_assets'))

@app.route('/fixed-assets/<int:id>/delete', methods=['POST'])
def delete_fixed_asset(id):
    fa = FixedAsset.query.get_or_404(id)
    db.session.delete(fa)
    db.session.commit()
    flash('Хөрөнгө устгагдлаа!', 'success')
    return redirect(url_for('fixed_assets'))

# ===================== VAT BATCH JOURNAL =====================

@app.route('/vat/batch-journal', methods=['GET', 'POST'])
def vat_batch_journal():
    from sqlalchemy import func as sqlfunc

    months_list = [m[0] for m in
        db.session.query(VatRecord.month).filter(VatRecord.month != None)
        .distinct().order_by(VatRecord.month).all()]

    accounts = Account.query.filter_by(is_active=True).order_by(Account.code).all()

    def find_by_code(code):
        for a in accounts:
            if a.code == code:
                return a.id
        return ''

    def find_acc(keyword):
        for a in accounts:
            if keyword.lower() in (a.code + ' ' + a.name).lower():
                return a.id
        return ''

    defaults = {
        'out_dt':  find_by_code('1210') or find_acc('авлага'),
        'out_kt1': find_by_code('6110') or find_acc('орлого'),
        'out_kt2': find_by_code('3132') or find_acc('нөат өглөг'),   # НӨАТ өглөг
        'in_dt1':  find_by_code('7200') or find_acc('тээвэр зардал'),
        'in_dt2':  find_by_code('1352') or find_acc('оруулсан нөат'),  # НӨАТ авлага
        'in_kt':   find_by_code('3110') or find_acc('нийлүүлэгч'),
    }

    # Monthly preview
    preview = []
    for m in months_list:
        rows = db.session.query(
            VatRecord.type,
            sqlfunc.count().label('cnt'),
            sqlfunc.sum(VatRecord.amount).label('amt'),
            sqlfunc.sum(VatRecord.vat_amount).label('vat'),
            sqlfunc.sum(VatRecord.total_amount).label('total')
        ).filter_by(month=m).group_by(VatRecord.type).all()
        out = next((r for r in rows if r[0]=='out'), None)
        inp = next((r for r in rows if r[0]=='in'),  None)
        # check if journal already created for this month
        out_exists = Journal.query.filter(Journal.number==f'EB-OUT-2026-{m:02d}').first()
        in_exists  = Journal.query.filter(Journal.number==f'EB-IN-2026-{m:02d}').first()
        preview.append(dict(
            month=m,
            out_cnt=out[1] if out else 0, out_amt=out[2] or 0 if out else 0,
            out_vat=out[3] or 0 if out else 0, out_total=out[4] or 0 if out else 0,
            in_cnt=inp[1] if inp else 0,  in_amt=inp[2] or 0 if inp else 0,
            in_vat=inp[3] or 0 if inp else 0,  in_total=inp[4] or 0 if inp else 0,
            out_done=bool(out_exists), in_done=bool(in_exists)
        ))

    if request.method == 'POST':
        sel_months = request.form.getlist('months')
        vtype      = request.form.get('vtype', 'both')
        year       = 2026

        def get_acc(field):
            v = request.form.get(field, '')
            return int(v) if v else None

        out_dt  = get_acc('out_dt');  out_kt1 = get_acc('out_kt1'); out_kt2 = get_acc('out_kt2')
        in_dt1  = get_acc('in_dt1'); in_dt2  = get_acc('in_dt2');  in_kt   = get_acc('in_kt')

        created = 0
        for m_str in sel_months:
            m = int(m_str)
            last_day = {1:31,2:28,3:31,4:30,5:31,6:30,7:31,8:31,9:30,10:31,11:30,12:31}.get(m,30)
            jdate = date(year, m, last_day)

            if vtype in ('out','both'):
                recs = VatRecord.query.filter_by(type='out', month=m).all()
                num  = f'EB-OUT-{year}-{m:02d}'
                if recs and not Journal.query.filter_by(number=num).first():
                    total_amt = round(sum(r.amount for r in recs), 2)
                    total_vat = round(sum(r.vat_amount for r in recs), 2)
                    j = Journal(date=jdate, number=num, status='posted',
                        description=f'eBarimt борлуулалт — {year}/{m:02d} ({len(recs)} баримт)',
                        reference=f'eBarimt {m}-р сар')
                    db.session.add(j); db.session.flush()
                    if total_amt:
                        db.session.add(JournalLine(journal_id=j.id,
                            debit_account_id=out_dt, credit_account_id=out_kt1,
                            amount=total_amt, description=f'Борлуулалтын орлого {m}-р сар'))
                    if total_vat:
                        db.session.add(JournalLine(journal_id=j.id,
                            debit_account_id=out_dt, credit_account_id=out_kt2,
                            amount=total_vat, description=f'НӨТ-ын өглөг {m}-р сар'))
                    created += 1

            if vtype in ('in','both'):
                recs = VatRecord.query.filter_by(type='in', month=m).all()
                num  = f'EB-IN-{year}-{m:02d}'
                if recs and not Journal.query.filter_by(number=num).first():
                    total_amt = round(sum(r.amount for r in recs), 2)
                    total_vat = round(sum(r.vat_amount for r in recs), 2)
                    j = Journal(date=jdate, number=num, status='posted',
                        description=f'eBarimt худалдан авалт — {year}/{m:02d} ({len(recs)} баримт)',
                        reference=f'eBarimt {m}-р сар')
                    db.session.add(j); db.session.flush()
                    if total_amt:
                        db.session.add(JournalLine(journal_id=j.id,
                            debit_account_id=in_dt1, credit_account_id=in_kt,
                            amount=total_amt, description=f'Худалдан авалт {m}-р сар'))
                    if total_vat:
                        db.session.add(JournalLine(journal_id=j.id,
                            debit_account_id=in_dt2, credit_account_id=in_kt,
                            amount=total_vat, description=f'НӨТ-ын авлага {m}-р сар'))
                    created += 1

        db.session.commit()
        flash(f'{created} журналын бичилт амжилттай үүслээ!', 'success')
        return redirect(url_for('vat'))

    return render_template('vat_batch_journal.html',
                           preview=preview, accounts=accounts,
                           defaults=defaults, months_list=months_list)

# ===================== EXCEL EXPORT =====================

def _xl_style(wb):
    """Return commonly used openpyxl styles."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    hdr_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    hdr_fill  = PatternFill('solid', fgColor='1A3C5E')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    num_fmt   = '#,##0'
    thin      = Side(style='thin', color='CCCCCC')
    border    = Border(left=thin, right=thin, top=thin, bottom=thin)
    return hdr_font, hdr_fill, hdr_align, num_fmt, border

def _xl_header(ws, cols):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    hdr_font, hdr_fill, hdr_align, _, border = _xl_style(None)
    for c_idx, (title, width) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c_idx, value=title)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
        ws.column_dimensions[ws.cell(row=1, column=c_idx).column_letter].width = width
    ws.row_dimensions[1].height = 22

def _xl_response(wb, filename):
    import io
    from flask import send_file
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return send_file(buf, download_name=filename,
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── НӨАТ бүртгэл ──────────────────────────────────────────────────────
@app.route('/export/vat')
def export_vat():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, numbers
    from sqlalchemy import func

    vtype  = request.args.get('type', '')
    month  = request.args.get('month', type=int)
    q = VatRecord.query
    if vtype:  q = q.filter_by(type=vtype)
    if month:  q = q.filter_by(month=month)
    records = q.order_by(VatRecord.date, VatRecord.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'НӨАТ бүртгэл'

    cols = [('Огноо',12),('Төрөл',8),('ДДТД',16),('Нэхэмжлэл №',14),
            ('Харилцагч',28),('Регистр',12),
            ('Суурь дүн',14),('НӨАТ',12),('Нийт дүн',14),
            ('Сар',5),('Эх үүсвэр',10)]
    _xl_header(ws, cols)

    from openpyxl.styles import Font, Border, Side, Alignment
    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_i, r in enumerate(records, 2):
        tp = 'Борлуулалт' if r.type == 'out' else 'Худалдан авалт'
        vals = [r.date, tp, r.ddtd or '', r.invoice_no or '',
                r.partner_name or '', r.partner_register or '',
                r.amount, r.vat_amount, r.total_amount,
                r.month, r.source or '']
        for c_i, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=c_i, value=v)
            cell.border = brd
            if c_i in (7, 8, 9):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            if c_i == 2:
                cell.font = Font(color='198754' if r.type=='out' else 'DC3545', bold=True, size=9)

    # Totals row
    tr = len(records) + 2
    ws.cell(tr, 1, 'НИЙТ').font = Font(bold=True)
    ws.cell(tr, 7, sum(r.amount     for r in records)).number_format = '#,##0'
    ws.cell(tr, 8, sum(r.vat_amount for r in records)).number_format = '#,##0'
    ws.cell(tr, 9, sum(r.total_amount for r in records)).number_format = '#,##0'
    for c in (7,8,9):
        ws.cell(tr, c).font = Font(bold=True)
    ws.freeze_panes = 'A2'

    fname = f'НӨАТ_бүртгэл_{date.today()}.xlsx'
    return _xl_response(wb, fname)

# ── Гүйлгээний бүртгэл (журнал) ──────────────────────────────────────
@app.route('/export/journals')
def export_journals():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

    journals_q = Journal.query.order_by(Journal.date, Journal.number).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Гүйлгээний бүртгэл'

    cols = [('№',8),('Огноо',12),('Журнал №',14),('Тайлбар',32),
            ('Лавлагаа',14),('Дүн',14),('Төлөв',10)]
    _xl_header(ws, cols)

    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, j in enumerate(journals_q, 1):
        total = sum(l.amount for l in j.lines if l.debit_account_id)
        st_color = '198754' if j.status == 'posted' else 'FF9800'
        st_label = 'Баталгаажсан' if j.status == 'posted' else 'Ноорог'
        vals = [i, j.date, j.number, j.description or '', j.reference or '', total, st_label]
        for c_i, v in enumerate(vals, 1):
            cell = ws.cell(row=i+1, column=c_i, value=v)
            cell.border = brd
            if c_i == 6:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            if c_i == 7:
                cell.font = Font(color=st_color, bold=True, size=9)
    ws.freeze_panes = 'A2'

    # Journal lines sheet
    ws2 = wb.create_sheet('Мөрүүд')
    cols2 = [('Журнал №',14),('Огноо',12),('ДТ код',10),('ДТ нэр',24),
             ('КТ код',10),('КТ нэр',24),('Дүн',14),('Тайлбар',28)]
    _xl_header(ws2, cols2)
    row2 = 2
    for j in journals_q:
        for l in j.lines:
            dt = Account.query.get(l.debit_account_id)  if l.debit_account_id  else None
            kt = Account.query.get(l.credit_account_id) if l.credit_account_id else None
            vals = [j.number, j.date,
                    dt.code if dt else '', dt.name if dt else '',
                    kt.code if kt else '', kt.name if kt else '',
                    l.amount, l.description or '']
            for c_i, v in enumerate(vals, 1):
                cell = ws2.cell(row=row2, column=c_i, value=v)
                cell.border = Border(left=Side(style='thin',color='CCCCCC'),right=Side(style='thin',color='CCCCCC'),top=Side(style='thin',color='CCCCCC'),bottom=Side(style='thin',color='CCCCCC'))
                if c_i == 7:
                    cell.number_format = '#,##0'
                    cell.alignment = Alignment(horizontal='right')
            row2 += 1
    ws2.freeze_panes = 'A2'

    return _xl_response(wb, f'Журнал_{date.today()}.xlsx')

# ── Үндсэн хөрөнгө ───────────────────────────────────────────────────
@app.route('/export/fixed-assets')
def export_fixed_assets():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment

    assets = FixedAsset.query.order_by(FixedAsset.code).all()
    wb = Workbook(); ws = wb.active; ws.title = 'Үндсэн хөрөнгө'

    cols = [('Код',10),('Нэр',28),('Ангилал',18),('Авсан огноо',13),
            ('Анхны өртөг',15),('Хуримт. элэгдэл',16),('Дансны үнэ',13),
            ('Ашиглалт (жил)',10),('Жилийн элэгдэл',15),('Сарын элэгдэл',14),
            ('Элэгдэл %',10),('Төлөв',10),('Байршил',18),('Хариуцагч',16)]
    _xl_header(ws, cols)
    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, a in enumerate(assets, 2):
        vals = [a.code or '', a.name, a.category,
                a.purchase_date, a.purchase_amount,
                a.accumulated_depreciation, a.book_value,
                a.useful_life, a.annual_depreciation, a.monthly_depreciation,
                round(a.depreciation_pct, 1),
                'Идэвхтэй' if a.status=='active' else 'Хасагдсан' if a.status=='disposed' else 'Борлуулсан',
                a.location or '', a.responsible_person or '']
        for c_i, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c_i, value=v)
            cell.border = brd
            if c_i in (5,6,7,9,10):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
            if c_i == 11:
                cell.number_format = '0.0"%"'
                cell.alignment = Alignment(horizontal='center')

    # Totals
    tr = len(assets) + 2
    ws.cell(tr, 1, 'НИЙТ').font = Font(bold=True)
    for c_i, attr in [(5,'purchase_amount'),(6,'accumulated_depreciation'),(7,'book_value')]:
        ws.cell(tr, c_i, sum(getattr(a, attr) for a in assets)).number_format = '#,##0'
        ws.cell(tr, c_i).font = Font(bold=True)
    ws.freeze_panes = 'A2'

    return _xl_response(wb, f'Үндсэн_хөрөнгө_{date.today()}.xlsx')

# ── Гүйлгээ баланс ───────────────────────────────────────────────────
@app.route('/export/trial-balance')
def export_trial_balance():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
    from sqlalchemy import func as sqlfunc

    accounts_all = Account.query.filter_by(is_active=True).order_by(Account.code).all()
    acc_map = {a.id: a for a in accounts_all}

    dt_sums = dict(db.session.query(JournalLine.debit_account_id,  sqlfunc.sum(JournalLine.amount)).group_by(JournalLine.debit_account_id).all())
    kt_sums = dict(db.session.query(JournalLine.credit_account_id, sqlfunc.sum(JournalLine.amount)).group_by(JournalLine.credit_account_id).all())

    wb = Workbook(); ws = wb.active; ws.title = 'Гүйлгээ баланс'
    cols = [('Дансны код',12),('Дансны нэр',32),('Дебет нийт',15),
            ('Кредит нийт',15),('Үлдэгдэл (ДТ)',15),('Үлдэгдэл (КТ)',15)]
    _xl_header(ws, cols)
    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_i = 2
    tot_dt = tot_kt = tot_bal_dt = tot_bal_kt = 0
    for a in accounts_all:
        dt = dt_sums.get(a.id, 0) or 0
        kt = kt_sums.get(a.id, 0) or 0
        if dt == 0 and kt == 0:
            continue
        bal = dt - kt
        bal_dt = bal if bal > 0 else 0
        bal_kt = -bal if bal < 0 else 0
        tot_dt += dt; tot_kt += kt
        tot_bal_dt += bal_dt; tot_bal_kt += bal_kt

        for c_i, v in enumerate([a.code, a.name, dt, kt, bal_dt, bal_kt], 1):
            cell = ws.cell(row=row_i, column=c_i, value=v)
            cell.border = brd
            if c_i >= 3:
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
        row_i += 1

    # Totals
    for c_i, v in enumerate(['', 'НИЙТ', tot_dt, tot_kt, tot_bal_dt, tot_bal_kt], 1):
        cell = ws.cell(row=row_i, column=c_i, value=v)
        cell.font = Font(bold=True)
        if c_i >= 3:
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right')
        cell.fill = PatternFill('solid', fgColor='E8F0FE')
    ws.freeze_panes = 'A2'

    return _xl_response(wb, f'Гүйлгээ_баланс_{date.today()}.xlsx')

# ── Банкны хуулга ────────────────────────────────────────────────────
@app.route('/export/cash')
def export_cash():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Border, Side, Alignment

    bank  = request.args.get('bank', '')
    month = request.args.get('month', type=int)
    q = CashTransaction.query
    if bank:  q = q.filter_by(bank=bank)
    if month: q = q.filter_by(month=month)
    txns = q.order_by(CashTransaction.date, CashTransaction.id).all()

    wb = Workbook(); ws = wb.active; ws.title = 'Банкны хуулга'
    cols = [('Огноо',12),('Банк',10),('Тайлбар',36),('Харилцагч',24),
            ('Орлого',14),('Зарлага',14),('Орлогын ангилал',16),('Зарлагын ангилал',16)]
    _xl_header(ws, cols)
    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, t in enumerate(txns, 2):
        partner = ''
        if t.partner_id:
            p = db.session.get(Partner, t.partner_id)
            partner = p.name if p else ''
        vals = [t.date, t.bank or '', t.description or '', partner,
                t.income or 0, t.expense or 0,
                t.income_cat or '', t.expense_cat or '']
        for c_i, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c_i, value=v)
            cell.border = brd
            if c_i in (5, 6):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')
                if c_i == 5 and (v or 0) > 0:
                    cell.font = Font(color='198754')
                if c_i == 6 and (v or 0) > 0:
                    cell.font = Font(color='DC3545')

    tr = len(txns) + 2
    ws.cell(tr, 1, 'НИЙТ').font = Font(bold=True)
    ws.cell(tr, 5, sum(t.income  or 0 for t in txns)).number_format = '#,##0'
    ws.cell(tr, 6, sum(t.expense or 0 for t in txns)).number_format = '#,##0'
    for c in (5, 6):
        ws.cell(tr, c).font = Font(bold=True)
    ws.freeze_panes = 'A2'

    return _xl_response(wb, f'Банкны_хуулга_{date.today()}.xlsx')

# ===================== FORMS (анхан шатны маягт) =====================

@app.route('/forms')
def forms():
    return render_template('forms.html')

@app.route('/forms/mh1')
def form_mh1():
    num = request.args.get('num', '')
    return render_template('form_mh1.html', num=num, today=date.today())

@app.route('/forms/mh2')
def form_mh2():
    num = request.args.get('num', '')
    return render_template('form_mh2.html', num=num, today=date.today())

@app.route('/forms/t1')
def form_t1():
    rec_id  = request.args.get('rec',  type=int)
    pay_id  = request.args.get('pay',  type=int)
    rec     = Receivable.query.get(rec_id) if rec_id else None
    pay     = Payable.query.get(pay_id)    if pay_id else None
    num     = request.args.get('num', '')
    return render_template('form_t1.html', rec=rec, pay=pay, num=num, today=date.today())

@app.route('/forms/t2')
def form_t2():
    num = request.args.get('num', '')
    return render_template('form_t2.html', num=num, today=date.today())

# ===================== SEED DATA =====================

def seed_accounts():
    if Account.query.count() > 0:
        return
    accounts = [
        # Хөрөнгө
        ('1000', 'Мөнгөн хөрөнгө', 'asset', None),
        ('1010', 'Касс', 'asset', '1000'),
        ('1020', 'Банкны данс', 'asset', '1000'),
        ('1100', 'Авлага', 'asset', None),
        ('1110', 'Худалдан авагчдаас авлага', 'asset', '1100'),
        ('1200', 'Бараа материал', 'asset', None),
        ('1210', 'Бараа', 'asset', '1200'),
        ('1500', 'Үндсэн хөрөнгө', 'asset', None),
        ('1510', 'Тоног төхөөрөмж', 'asset', '1500'),
        ('1590', 'Хуримтлагдсан элэгдэл', 'asset', '1500'),
        # Өр төлбөр
        ('2000', 'Богино хугацааны өр', 'liability', None),
        ('2010', 'Нийлүүлэгчдэд өглөг', 'liability', '2000'),
        ('2020', 'Татварын өглөг', 'liability', '2000'),
        ('2021', 'НӨАТ өглөг', 'liability', '2020'),
        ('2022', 'ХХОАТ өглөг', 'liability', '2020'),
        ('2030', 'Цалингийн өглөг', 'liability', '2000'),
        ('2500', 'Урт хугацааны өр', 'liability', None),
        ('2510', 'Зээл', 'liability', '2500'),
        # Өмч
        ('3000', 'Өмч', 'equity', None),
        ('3010', 'Дүрмийн сан', 'equity', '3000'),
        ('3020', 'Хуримтлагдсан ашиг', 'equity', '3000'),
        # Орлого
        ('4000', 'Борлуулалтын орлого', 'income', None),
        ('4010', 'Бараа борлуулалт', 'income', '4000'),
        ('4020', 'Үйлчилгээний орлого', 'income', '4000'),
        ('4500', 'Бусад орлого', 'income', None),
        # Зардал
        ('5000', 'Борлуулалтын өртөг', 'expense', None),
        ('5010', 'Бараа материалын зардал', 'expense', '5000'),
        ('5100', 'Цалин хөлс', 'expense', None),
        ('5110', 'Үндсэн цалин', 'expense', '5100'),
        ('5120', 'НДШ зардал (ажил олгогч)', 'expense', '5100'),
        ('5200', 'Үйл ажиллагааны зардал', 'expense', None),
        ('5210', 'Түрээс', 'expense', '5200'),
        ('5220', 'Цахилгаан, дулаан', 'expense', '5200'),
        ('5230', 'Утас, интернет', 'expense', '5200'),
        ('5240', 'Тээвэр зардал', 'expense', '5200'),
        ('5300', 'Элэгдлийн зардал', 'expense', None),
        ('5900', 'Бусад зардал', 'expense', None),
    ]
    code_to_id = {}
    for code, name, type_, parent_code in accounts:
        parent_id = code_to_id.get(parent_code) if parent_code else None
        acc = Account(code=code, name=name, type=type_, parent_id=parent_id)
        db.session.add(acc)
        db.session.flush()
        code_to_id[code] = acc.id
    db.session.commit()

with app.app_context():
    db.create_all()
    seed_accounts()
    seed_employees()

# ═══════════════════════════════════════════════════════════════
#  ЦАЛИНГИЙН БҮРТГЭЛ
# ═══════════════════════════════════════════════════════════════

@app.route('/salary')
def salary():
    month = request.args.get('month', datetime.now().month, type=int)
    year  = request.args.get('year',  datetime.now().year,  type=int)

    emps = Employee.query.filter_by(company='ТҮМЭН ТЭЭХ ХХК', is_active=True).order_by(Employee.id).all()

    total_hrs = MONTH_HOURS.get(month, 176)
    rows = []
    sum_niit = sum_emndsh = sum_emndsh_org = sum_hhoat = sum_adv = sum_gart = 0.0

    for emp in emps:
        # Тухайн сарын бүртгэл хайх
        rec = SalaryRecord.query.filter_by(employee_id=emp.id, year=year, month=month).first()

        if rec:
            worked = rec.worked_hrs
            sales  = rec.sales_bonus
            leave  = rec.leave_pay
            adv_ov = rec.adv_override
        else:
            worked = total_hrs   # Стандарт: бүтэн сар
            sales  = 0.0
            leave  = 0.0
            adv_ov = None

        r = calc_salary(emp.base_salary, worked, month,
                        phone=emp.phone_allowance,
                        sales=sales, leave=leave,
                        adv_override=adv_ov if adv_ov is not None else emp.adv_base)

        sum_niit       += r['niit']
        sum_emndsh     += r['emndsh']
        sum_emndsh_org += r['emndsh_org']
        sum_hhoat      += r['hhoat']
        sum_adv        += r['adv']
        sum_gart       += r['gart']

        rows.append({
            'emp':    emp,
            'worked': worked,
            'sales':  sales,
            'leave':  leave,
            'rec':    rec,
            **r,
        })

    months = list(range(1, 13))
    return render_template('salary.html',
        rows=rows, month=month, year=year, months=months,
        total_hrs=total_hrs,
        sum_niit=sum_niit, sum_emndsh=sum_emndsh,
        sum_emndsh_org=sum_emndsh_org,
        sum_hhoat=sum_hhoat, sum_adv=sum_adv, sum_gart=sum_gart,
    )


@app.route('/salary/update', methods=['POST'])
def salary_update():
    """Ажилтны цалингийн мэдээлэл шинэчлэх (AJAX)"""
    data = request.get_json() or {}
    emp_id = data.get('emp_id', type=int) if isinstance(data.get('emp_id'), int) else int(data.get('emp_id', 0))
    year   = int(data.get('year',  datetime.now().year))
    month  = int(data.get('month', datetime.now().month))
    field  = data.get('field', '')
    value  = data.get('value')

    emp = db.session.get(Employee, emp_id)
    if not emp:
        return jsonify({'ok': False, 'error': 'Ажилтан олдсонгүй'}), 404

    rec = SalaryRecord.query.filter_by(employee_id=emp_id, year=year, month=month).first()
    if not rec:
        total_hrs = MONTH_HOURS.get(month, 176)
        rec = SalaryRecord(
            employee_id=emp_id, year=year, month=month,
            base_salary=emp.base_salary, worked_hrs=total_hrs,
            total_hrs=total_hrs, phone=emp.phone_allowance,
        )
        db.session.add(rec)

    try:
        v = float(value) if value not in (None, '', 'null') else None
    except (ValueError, TypeError):
        v = None

    if field == 'worked':   rec.worked_hrs   = v if v is not None else MONTH_HOURS.get(month, 176)
    elif field == 'sales':  rec.sales_bonus  = v or 0
    elif field == 'leave':  rec.leave_pay    = v or 0
    elif field == 'adv':    rec.adv_override = v   # None = автомат (base*0.4 эсвэл adv_base)
    else:
        return jsonify({'ok': False, 'error': 'Буруу талбар'}), 400

    # Тооцоолол шинэчлэх
    adv_ov = rec.adv_override if rec.adv_override is not None else emp.adv_base
    r = calc_salary(emp.base_salary, rec.worked_hrs, month,
                    phone=emp.phone_allowance,
                    sales=rec.sales_bonus, leave=rec.leave_pay,
                    adv_override=adv_ov)
    rec.bod_salary   = r['bod']
    rec.total_income = r['niit']
    rec.emndsh       = r['emndsh']
    rec.hhoat_ded    = r['ded23']
    rec.hhoat        = r['hhoat']
    rec.advance      = r['adv']
    rec.net_pay      = r['gart']
    rec.updated_at   = datetime.utcnow()

    db.session.commit()
    return jsonify({'ok': True, **r, 'emndsh_org': r['emndsh_org']})


@app.route('/salary/employees')
def salary_employees():
    emps = Employee.query.filter_by(company='ТҮМЭН ТЭЭХ ХХК').order_by(Employee.id).all()
    return render_template('salary_employees.html', emps=emps)


@app.route('/salary/employees/<int:emp_id>/edit', methods=['GET', 'POST'])
def salary_emp_edit(emp_id):
    emp = db.session.get(Employee, emp_id)
    if not emp:
        flash('Ажилтан олдсонгүй', 'danger')
        return redirect(url_for('salary_employees'))
    if request.method == 'POST':
        emp.name            = request.form.get('name', emp.name).strip()
        emp.lastname        = request.form.get('lastname', '').strip()
        emp.tin             = request.form.get('tin', '').strip()
        emp.title           = request.form.get('title', '').strip()
        emp.base_salary     = float(request.form.get('base_salary', 0) or 0)
        emp.adv_base        = float(request.form.get('adv_base', 0) or 0) or None
        emp.phone_allowance = float(request.form.get('phone_allowance', 0) or 0)
        emp.phone           = request.form.get('phone', '').strip()
        emp.email           = request.form.get('email', '').strip()
        emp.bank            = request.form.get('bank', '').strip()
        emp.account         = request.form.get('account', '').strip()
        emp.exp_years       = int(request.form.get('exp_years', 0) or 0)
        db.session.commit()
        flash(f'{emp.full_name}-ийн мэдээлэл шинэчлэгдлээ.', 'success')
        return redirect(url_for('salary_employees'))
    return render_template('salary_emp_form.html', emp=emp)


if __name__ == '__main__':
    _port = int(_os.environ.get('FLASK_RUN_PORT', '5000'))
    app.run(debug=True, host='0.0.0.0', port=_port)
